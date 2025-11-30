#!/usr/bin/env python3
"""
Crea un report Excel dai saldi mensili estratti da Odoo
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def load_json_data(json_path):
    """Carica i dati dal file JSON"""
    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def create_excel_report(data, output_path):
    """Crea un report Excel dai dati"""
    print(f"\n{'='*80}")
    print("CREAZIONE REPORT EXCEL")
    print(f"{'='*80}\n")

    # Crea workbook
    wb = openpyxl.Workbook()

    # Rimuovi il foglio default
    wb.remove(wb.active)

    # FOGLIO 1: Saldi Mensili (una riga per conto, colonne per mesi)
    ws_balances = wb.create_sheet("Saldi Mensili 2024")
    create_balances_sheet(ws_balances, data)

    # FOGLIO 2: Dettaglio Movimenti (una riga per conto per mese)
    ws_details = wb.create_sheet("Dettaglio Mensile")
    create_details_sheet(ws_details, data)

    # FOGLIO 3: Riepilogo
    ws_summary = wb.create_sheet("Riepilogo")
    create_summary_sheet(ws_summary, data)

    # FOGLIO 4: Grafici di tendenza
    ws_trends = wb.create_sheet("Tendenze")
    create_trends_sheet(ws_trends, data)

    # Salva il file
    wb.save(output_path)
    print(f"\n[OK] Report Excel salvato: {output_path}\n")

def create_balances_sheet(ws, data):
    """Crea il foglio con i saldi mensili"""
    print("Creazione foglio: Saldi Mensili 2024...")

    # Stili
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header principale
    ws['A1'] = "SALDI MENSILI CONTI BANCARI - 2024"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:N1')

    # Info estrazione
    ws['A2'] = f"Estratto il: {datetime.fromisoformat(data['extraction_date']).strftime('%d/%m/%Y %H:%M')}"
    ws.merge_cells('A2:N2')

    # Header colonne
    row = 4
    headers = ['Codice', 'Nome Conto', 'Gen', 'Feb', 'Mar', 'Apr', 'Mag', 'Giu',
               'Lug', 'Ago', 'Set', 'Ott', 'Nov', 'Dic']

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Dati
    row = 5
    for account in data['accounts']:
        # Codice conto
        cell = ws.cell(row=row, column=1)
        cell.value = account['code']
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

        # Nome conto
        cell = ws.cell(row=row, column=2)
        cell.value = account['name']
        cell.border = border

        # Saldi mensili
        col = 3
        for month_end in [
            "2024-01-31", "2024-02-29", "2024-03-31", "2024-04-30",
            "2024-05-31", "2024-06-30", "2024-07-31", "2024-08-31",
            "2024-09-30", "2024-10-31", "2024-11-30", "2024-12-31"
        ]:
            balance = account['monthly_balances'].get(month_end, {}).get('balance', 0)
            cell = ws.cell(row=row, column=col)
            cell.value = balance
            cell.number_format = '#,##0.00'
            cell.border = border
            cell.alignment = Alignment(horizontal='right')

            # Colora in rosso i negativi
            if balance < 0:
                cell.font = Font(color="FF0000")

            col += 1

        row += 1

    # Riga TOTALE
    cell = ws.cell(row=row, column=1)
    cell.value = "TOTALE"
    cell.font = Font(bold=True)
    cell.border = border
    ws.merge_cells(f'A{row}:B{row}')

    col = 3
    for month_end in [
        "2024-01-31", "2024-02-29", "2024-03-31", "2024-04-30",
        "2024-05-31", "2024-06-30", "2024-07-31", "2024-08-31",
        "2024-09-30", "2024-10-31", "2024-11-30", "2024-12-31"
    ]:
        total = sum(
            account['monthly_balances'].get(month_end, {}).get('balance', 0)
            for account in data['accounts']
        )
        cell = ws.cell(row=row, column=col)
        cell.value = total
        cell.number_format = '#,##0.00'
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal='right')

        if total < 0:
            cell.font = Font(color="FF0000", bold=True)

        col += 1

    # Adatta larghezza colonne
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40
    for col in range(3, 15):
        ws.column_dimensions[get_column_letter(col)].width = 14

def create_details_sheet(ws, data):
    """Crea il foglio con i dettagli mensili"""
    print("Creazione foglio: Dettaglio Mensile...")

    # Stili
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header principale
    ws['A1'] = "DETTAGLIO MENSILE - 2024"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')

    # Header colonne
    row = 3
    headers = ['Codice', 'Nome Conto', 'Mese', 'Saldo', 'Dare', 'Avere', 'N. Movimenti']

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Dati
    row = 4
    month_names = {
        "2024-01-31": "Gennaio",
        "2024-02-29": "Febbraio",
        "2024-03-31": "Marzo",
        "2024-04-30": "Aprile",
        "2024-05-31": "Maggio",
        "2024-06-30": "Giugno",
        "2024-07-31": "Luglio",
        "2024-08-31": "Agosto",
        "2024-09-30": "Settembre",
        "2024-10-31": "Ottobre",
        "2024-11-30": "Novembre",
        "2024-12-31": "Dicembre"
    }

    for account in data['accounts']:
        for month_end, month_name in month_names.items():
            balance_data = account['monthly_balances'].get(month_end, {})

            # Codice
            cell = ws.cell(row=row, column=1)
            cell.value = account['code']
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            # Nome
            cell = ws.cell(row=row, column=2)
            cell.value = account['name']
            cell.border = border

            # Mese
            cell = ws.cell(row=row, column=3)
            cell.value = month_name
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            # Saldo
            cell = ws.cell(row=row, column=4)
            cell.value = balance_data.get('balance', 0)
            cell.number_format = '#,##0.00'
            cell.border = border
            cell.alignment = Alignment(horizontal='right')
            if balance_data.get('balance', 0) < 0:
                cell.font = Font(color="FF0000")

            # Dare
            cell = ws.cell(row=row, column=5)
            cell.value = balance_data.get('debit', 0)
            cell.number_format = '#,##0.00'
            cell.border = border
            cell.alignment = Alignment(horizontal='right')

            # Avere
            cell = ws.cell(row=row, column=6)
            cell.value = balance_data.get('credit', 0)
            cell.number_format = '#,##0.00'
            cell.border = border
            cell.alignment = Alignment(horizontal='right')

            # Movimenti
            cell = ws.cell(row=row, column=7)
            cell.value = balance_data.get('movements', 0)
            cell.number_format = '#,##0'
            cell.border = border
            cell.alignment = Alignment(horizontal='right')

            row += 1

    # Adatta larghezza colonne
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14

def create_summary_sheet(ws, data):
    """Crea il foglio riepilogativo"""
    print("Creazione foglio: Riepilogo...")

    # Header
    ws['A1'] = "RIEPILOGO SALDI BANCARI 2024"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')

    ws['A2'] = f"Estratto il: {datetime.fromisoformat(data['extraction_date']).strftime('%d/%m/%Y %H:%M')}"
    ws.merge_cells('A2:D2')

    # Info generali
    row = 4
    ws[f'A{row}'] = "Numero conti analizzati:"
    ws[f'B{row}'] = len(data['accounts'])
    ws[f'A{row}'].font = Font(bold=True)

    row += 1
    ws[f'A{row}'] = "Periodo:"
    ws[f'B{row}'] = "01/01/2024 - 31/12/2024"
    ws[f'A{row}'].font = Font(bold=True)

    # Saldo finale (31/12/2024)
    row += 3
    ws[f'A{row}'] = "SALDO TOTALE AL 31/12/2024"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    ws.merge_cells(f'A{row}:B{row}')

    total_balance = sum(
        account['monthly_balances'].get('2024-12-31', {}).get('balance', 0)
        for account in data['accounts']
    )

    row += 1
    ws[f'A{row}'] = "Totale:"
    ws[f'B{row}'] = total_balance
    ws[f'B{row}'].number_format = '#,##0.00 "CHF"'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True, size=14, color="0000FF" if total_balance >= 0 else "FF0000")

    # Dettaglio per conto
    row += 3
    ws[f'A{row}'] = "DETTAGLIO PER CONTO (31/12/2024)"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws.merge_cells(f'A{row}:D{row}')

    # Header tabella
    row += 1
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws[f'A{row}'] = "Codice"
    ws[f'B{row}'] = "Nome Conto"
    ws[f'C{row}'] = "Tipo"
    ws[f'D{row}'] = "Saldo"

    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row}'].fill = header_fill
        ws[f'{col}{row}'].font = header_font
        ws[f'{col}{row}'].alignment = Alignment(horizontal='center')

    # Dati conti
    row += 1
    for account in sorted(data['accounts'], key=lambda x: x['monthly_balances'].get('2024-12-31', {}).get('balance', 0), reverse=True):
        balance = account['monthly_balances'].get('2024-12-31', {}).get('balance', 0)

        ws[f'A{row}'] = account['code']
        ws[f'B{row}'] = account['name']
        ws[f'C{row}'] = account.get('account_type', 'N/A')
        ws[f'D{row}'] = balance
        ws[f'D{row}'].number_format = '#,##0.00'

        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        ws[f'D{row}'].alignment = Alignment(horizontal='right')

        if balance < 0:
            ws[f'D{row}'].font = Font(color="FF0000")

        row += 1

    # Adatta larghezza colonne
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

def create_trends_sheet(ws, data):
    """Crea il foglio con le tendenze"""
    print("Creazione foglio: Tendenze...")

    # Header
    ws['A1'] = "TENDENZE MENSILI - 2024"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:B1')

    # Calcola saldo totale per ogni mese
    row = 3
    ws[f'A{row}'] = "Mese"
    ws[f'B{row}'] = "Saldo Totale"

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'].fill = header_fill
    ws[f'B{row}'].font = header_font
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    ws[f'B{row}'].alignment = Alignment(horizontal='center')

    month_names = {
        "2024-01-31": "Gennaio",
        "2024-02-29": "Febbraio",
        "2024-03-31": "Marzo",
        "2024-04-30": "Aprile",
        "2024-05-31": "Maggio",
        "2024-06-30": "Giugno",
        "2024-07-31": "Luglio",
        "2024-08-31": "Agosto",
        "2024-09-30": "Settembre",
        "2024-10-31": "Ottobre",
        "2024-11-30": "Novembre",
        "2024-12-31": "Dicembre"
    }

    row = 4
    for month_end, month_name in month_names.items():
        total = sum(
            account['monthly_balances'].get(month_end, {}).get('balance', 0)
            for account in data['accounts']
        )

        ws[f'A{row}'] = month_name
        ws[f'B{row}'] = total
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        ws[f'B{row}'].alignment = Alignment(horizontal='right')

        if total < 0:
            ws[f'B{row}'].font = Font(color="FF0000")

        row += 1

    # Adatta larghezza colonne
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20

def main():
    """Funzione principale"""
    try:
        # Percorsi file
        json_path = r"C:\Users\lapa\Desktop\Claude Code\app-hub-platform\data-estratti\ODOO-BALANCES-2024-CLEAN.json"
        excel_path = r"C:\Users\lapa\Desktop\Claude Code\app-hub-platform\data-estratti\ODOO-BALANCES-2024-REPORT.xlsx"

        # Carica dati
        print("\nCaricamento dati da JSON...")
        data = load_json_data(json_path)

        # Crea report
        create_excel_report(data, excel_path)

        print("\n[OK] Report Excel creato con successo!\n")
        return 0

    except Exception as e:
        print(f"\n[ERRORE]: {str(e)}")
        import traceback
        traceback.print_exc()
        return 1

if __name__ == "__main__":
    exit(main())
