#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Genera Excel Report Chiusura 2024
Per Commercialista Patrick Angstmann
"""

import sys
import io
import json
from datetime import datetime

# Fix encoding su Windows
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("⚠️  openpyxl non installato. Installa con: pip install openpyxl")

class ExcelReportGenerator:
    def __init__(self, json_file):
        with open(json_file, 'r', encoding='utf-8') as f:
            self.data = json.load(f)

        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Rimuovi foglio default

        # Stili
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=12)
        self.title_font = Font(bold=True, size=14)
        self.warning_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        self.error_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        self.ok_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def create_summary_sheet(self):
        """Sheet 1: Summary"""
        ws = self.wb.create_sheet("1. Summary")

        # Intestazione
        ws['A1'] = "REPORT CHIUSURA BILANCIO 2024"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')

        ws['A2'] = "Lapa Delikatessen"
        ws['A2'].font = Font(bold=True, size=14)
        ws.merge_cells('A2:F2')

        ws['A3'] = f"Generato il: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws.merge_cells('A3:F3')

        ws['A4'] = "Periodo: 01.01.2024 - 31.12.2024"
        ws.merge_cells('A4:F4')

        ws['A5'] = "Commercialista: Patrick Angstmann - PAGG Treuhand AG (p.angstmann@pagg.ch)"
        ws.merge_cells('A5:F5')

        # Balance Sheet
        row = 7
        ws[f'A{row}'] = "BALANCE SHEET al 31.12.2024"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:F{row}')

        row += 1
        bs_totals = self.data['balance_sheet']['totals']

        headers = ['Voce', 'Importo CHF', 'Note']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border

        row += 1
        ws[f'A{row}'] = "ASSETS"
        ws[f'B{row}'] = bs_totals['assets']
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'C{row}'] = "Attivo"

        row += 1
        ws[f'A{row}'] = "LIABILITIES"
        ws[f'B{row}'] = bs_totals['liabilities']
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'C{row}'] = "Passivo"

        row += 1
        ws[f'A{row}'] = "EQUITY"
        ws[f'B{row}'] = bs_totals['equity']
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'C{row}'] = "Patrimonio Netto"

        row += 1
        ws[f'A{row}'] = "BALANCE CHECK"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = bs_totals['balance_check']
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'B{row}'].font = Font(bold=True)

        if abs(bs_totals['balance_check']) > 0.01:
            ws[f'B{row}'].fill = self.error_fill
            ws[f'C{row}'] = "❌ NON BILANCIATO"
            ws[f'C{row}'].fill = self.error_fill
        else:
            ws[f'B{row}'].fill = self.ok_fill
            ws[f'C{row}'] = "✓ OK"
            ws[f'C{row}'].fill = self.ok_fill

        # Profit & Loss
        row += 3
        ws[f'A{row}'] = "PROFIT & LOSS 2024"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:F{row}')

        row += 1
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border

        pl_totals = self.data['profit_loss']['totals']

        row += 1
        ws[f'A{row}'] = "INCOME"
        ws[f'B{row}'] = pl_totals['income']
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'C{row}'] = "Ricavi"

        row += 1
        ws[f'A{row}'] = "EXPENSES"
        ws[f'B{row}'] = pl_totals['expenses']
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'C{row}'] = "Costi"

        row += 1
        ws[f'A{row}'] = "NET PROFIT"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = pl_totals['net_profit']
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'] = "Utile Netto"

        # Checklist
        row += 3
        ws[f'A{row}'] = "CHECKLIST VALIDAZIONE"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:F{row}')

        row += 1
        headers = ['Konto', 'Nome', 'Saldo CHF', 'Status', 'Note']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border

        for code, data in self.data['checklist_validation'].items():
            row += 1
            ws[f'A{row}'] = data['code']
            ws[f'B{row}'] = data['name']
            ws[f'C{row}'] = data.get('balance', 0.0)
            ws[f'C{row}'].number_format = '#,##0.00'
            ws[f'D{row}'] = data['status']
            ws[f'E{row}'] = ', '.join(data.get('issues', []))

            # Colora in base allo status
            if '✗' in data['status']:
                ws[f'D{row}'].fill = self.error_fill
                ws[f'D{row}'].font = Font(color="FFFFFF", bold=True)
            elif '⚠️' in data['status']:
                ws[f'D{row}'].fill = self.warning_fill
                ws[f'D{row}'].font = Font(bold=True)
            else:
                ws[f'D{row}'].fill = self.ok_fill
                ws[f'D{row}'].font = Font(color="FFFFFF", bold=True)

        # Auto-width
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20

        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['E'].width = 40

    def create_balance_sheet(self):
        """Sheet 2: Balance Sheet Dettagliato"""
        ws = self.wb.create_sheet("2. Balance Sheet")

        ws['A1'] = "BALANCE SHEET - Dettaglio Conti"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:E1')

        row = 3
        headers = ['Tipo Conto', 'Codice', 'Nome Conto', 'Saldo CHF']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border

        accounts_by_type = self.data['balance_sheet'].get('accounts_by_type', {})

        for account_type, accounts in sorted(accounts_by_type.items()):
            row += 1
            ws[f'A{row}'] = account_type.upper()
            ws[f'A{row}'].font = Font(bold=True, size=11)
            ws.merge_cells(f'A{row}:E{row}')
            ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

            subtotal = 0
            for acc in sorted(accounts, key=lambda x: x['code']):
                row += 1
                ws[f'A{row}'] = ""
                ws[f'B{row}'] = acc['code']
                ws[f'C{row}'] = acc['name']
                ws[f'D{row}'] = acc['balance']
                ws[f'D{row}'].number_format = '#,##0.00'
                subtotal += acc['balance']

            row += 1
            ws[f'C{row}'] = f"Subtotale {account_type}"
            ws[f'C{row}'].font = Font(bold=True)
            ws[f'D{row}'] = subtotal
            ws[f'D{row}'].number_format = '#,##0.00'
            ws[f'D{row}'].font = Font(bold=True)
            ws[f'D{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

        # Auto-width
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 18

    def create_profit_loss(self):
        """Sheet 3: Profit & Loss Dettagliato"""
        ws = self.wb.create_sheet("3. Profit & Loss")

        ws['A1'] = "PROFIT & LOSS 2024 - Dettaglio Conti"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:E1')

        row = 3
        headers = ['Tipo Conto', 'Codice', 'Nome Conto', 'Saldo CHF']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border

        accounts_by_type = self.data['profit_loss'].get('accounts_by_type', {})

        for account_type, accounts in sorted(accounts_by_type.items()):
            row += 1
            ws[f'A{row}'] = account_type.upper()
            ws[f'A{row}'].font = Font(bold=True, size=11)
            ws.merge_cells(f'A{row}:E{row}')
            ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

            subtotal = 0
            for acc in sorted(accounts, key=lambda x: x['code']):
                row += 1
                ws[f'A{row}'] = ""
                ws[f'B{row}'] = acc['code']
                ws[f'C{row}'] = acc['name']
                ws[f'D{row}'] = acc['balance']
                ws[f'D{row}'].number_format = '#,##0.00'
                subtotal += acc['balance']

            row += 1
            ws[f'C{row}'] = f"Subtotale {account_type}"
            ws[f'C{row}'].font = Font(bold=True)
            ws[f'D{row}'] = subtotal
            ws[f'D{row}'].number_format = '#,##0.00'
            ws[f'D{row}'].font = Font(bold=True)
            ws[f'D{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

        # Auto-width
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 18

    def create_chart_of_accounts(self):
        """Sheet 4: Piano dei Conti Completo"""
        ws = self.wb.create_sheet("4. Piano dei Conti")

        ws['A1'] = "PIANO DEI CONTI - Tutti i Conti Attivi"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')

        row = 3
        headers = ['Codice', 'Nome Conto', 'Tipo', 'Saldo CHF', 'Riconciliabile']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border

        for acc in self.data['accounts']:
            row += 1
            ws[f'A{row}'] = acc['code']
            ws[f'B{row}'] = acc['name']
            ws[f'C{row}'] = acc['type']
            ws[f'D{row}'] = acc['balance']
            ws[f'D{row}'].number_format = '#,##0.00'
            ws[f'E{row}'] = "Sì" if acc['reconcile'] else "No"

        # Auto-width
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15

        # Freeze panes
        ws.freeze_panes = 'A4'

    def create_bank_reconciliations(self):
        """Sheet 5: Riconciliazioni Bancarie"""
        ws = self.wb.create_sheet("5. Riconciliazioni Bancarie")

        ws['A1'] = "RICONCILIAZIONI BANCARIE al 31.12.2024"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')

        row = 3
        headers = ['Codice', 'Nome Conto', 'Saldo CHF', 'Movimenti Non Riconc.', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border

        for bank in self.data['bank_reconciliations']:
            row += 1
            ws[f'A{row}'] = bank['account_code']
            ws[f'B{row}'] = bank['account_name']
            ws[f'C{row}'] = bank['balance']
            ws[f'C{row}'].number_format = '#,##0.00'
            ws[f'D{row}'] = bank['unreconciled_count']

            if bank['unreconciled_count'] == 0:
                ws[f'E{row}'] = "✓ OK"
                ws[f'E{row}'].fill = self.ok_fill
                ws[f'E{row}'].font = Font(color="FFFFFF", bold=True)
            elif bank['unreconciled_count'] < 50:
                ws[f'E{row}'] = "⚠️ WARNING"
                ws[f'E{row}'].fill = self.warning_fill
            else:
                ws[f'E{row}'] = "❌ CRITICO"
                ws[f'E{row}'].fill = self.error_fill
                ws[f'E{row}'].font = Font(color="FFFFFF", bold=True)

        # Auto-width
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 15

    def create_vat_summary(self):
        """Sheet 6: Riepilogo IVA"""
        ws = self.wb.create_sheet("6. Riepilogo IVA")

        ws['A1'] = "RIEPILOGO IVA 2024"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:E1')

        row = 3
        headers = ['Codice', 'Nome Conto', 'Descrizione', 'Saldo CHF']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border

        for code, vat in self.data['vat_accounts'].items():
            row += 1
            ws[f'A{row}'] = vat['code']
            ws[f'B{row}'] = vat['name']
            ws[f'C{row}'] = vat['description']
            ws[f'D{row}'] = vat['balance']
            ws[f'D{row}'].number_format = '#,##0.00'

        row += 2
        ws[f'A{row}'] = "NOTE:"
        ws[f'A{row}'].font = Font(bold=True)
        ws.merge_cells(f'A{row}:E{row}')

        row += 1
        ws[f'A{row}'] = "- Verificare che i saldi IVA corrispondano alle dichiarazioni IVA 2024"
        ws.merge_cells(f'A{row}:E{row}')

        row += 1
        ws[f'A{row}'] = "- Konto 1170 (Vorsteuer): IVA versata su acquisti - credito verso Fisco"
        ws.merge_cells(f'A{row}:E{row}')

        row += 1
        ws[f'A{row}'] = "- Konto 2016 (Kreditor): IVA incassata su vendite - debito verso Fisco"
        ws.merge_cells(f'A{row}:E{row}')

        # Auto-width
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 18

    def generate(self, output_file):
        """Genera Excel completo"""
        print("Generazione Excel Report Chiusura 2024...")

        self.create_summary_sheet()
        print("  ✓ Sheet 1: Summary")

        self.create_balance_sheet()
        print("  ✓ Sheet 2: Balance Sheet")

        self.create_profit_loss()
        print("  ✓ Sheet 3: Profit & Loss")

        self.create_chart_of_accounts()
        print("  ✓ Sheet 4: Piano dei Conti")

        self.create_bank_reconciliations()
        print("  ✓ Sheet 5: Riconciliazioni Bancarie")

        self.create_vat_summary()
        print("  ✓ Sheet 6: Riepilogo IVA")

        self.wb.save(output_file)
        print(f"\n✓ Excel salvato: {output_file}")

def main():
    if not HAS_OPENPYXL:
        print("\n❌ Impossibile generare Excel senza openpyxl")
        print("   Installa con: pip install openpyxl")
        return 1

    try:
        json_file = "report-chiusura-2024.json"
        output_file = "REPORT-CHIUSURA-2024.xlsx"

        generator = ExcelReportGenerator(json_file)
        generator.generate(output_file)

        print("\n✓ Report Excel completato con successo!")
        return 0

    except Exception as e:
        print(f"\n✗ ERRORE: {e}")
        import traceback
        traceback.print_exc()
        return 1

if __name__ == "__main__":
    exit(main())
