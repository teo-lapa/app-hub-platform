#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ANALISI DETTAGLIATA 603 ANOMALIE FORNITORI
Esporta lista completa pagamenti senza fattura per azione correttiva
"""

import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# Carica dati JSON
print("Caricamento dati riconciliazione...")
with open('c:\\Users\\lapa\\Desktop\\Claude Code\\app-hub-platform\\riconciliazione-clienti-fornitori-2024.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# Estrai anomalie fornitori
print("\nEstrazione anomalie fornitori...")
anomalie_fornitori = []

for fornitore in data['top_fornitori_analysis']:
    fornitore_id = fornitore['partner_id']
    fornitore_name = fornitore['partner_name']
    balance = fornitore['balance']

    analysis = fornitore['analysis']
    discrepanze = analysis['discrepanze']

    for disc in discrepanze:
        anomalie_fornitori.append({
            'fornitore_id': fornitore_id,
            'fornitore_name': fornitore_name,
            'balance_fornitore': balance,
            'tipo_anomalia': disc['type'],
            'documento': disc['document'],
            'data': disc['date'],
            'importo': disc['amount'],
            'residuo': disc['residual'],
            'status': disc['status']
        })

print(f"Trovate {len(anomalie_fornitori)} anomalie fornitori")

# Crea DataFrame
df = pd.DataFrame(anomalie_fornitori)

# Statistiche
print("\n" + "="*80)
print("STATISTICHE ANOMALIE FORNITORI")
print("="*80)

print(f"\nTotale anomalie: {len(df)}")
print(f"Totale valore anomalie: CHF {df['importo'].sum():,.2f}")

print("\nDistribuzione per tipo anomalia:")
tipo_dist = df.groupby('tipo_anomalia').agg({
    'documento': 'count',
    'importo': 'sum'
}).rename(columns={'documento': 'count'})
print(tipo_dist)

print("\nTOP 10 Fornitori per numero anomalie:")
top_count = df.groupby('fornitore_name').size().sort_values(ascending=False).head(10)
print(top_count)

print("\nTOP 10 Fornitori per valore anomalie:")
top_value = df.groupby('fornitore_name')['importo'].sum().sort_values(ascending=False).head(10)
print(top_value)

# Crea Excel dettagliato
print("\nGenerazione Excel dettagliato...")
wb = Workbook()
ws = wb.active
ws.title = "Anomalie Fornitori"

# Header
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

headers = [
    'Fornitore ID', 'Fornitore', 'Saldo Fornitore', 'Tipo Anomalia',
    'Documento', 'Data', 'Importo CHF', 'Residuo CHF', 'Status'
]
ws.append(headers)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Dati
for _, row in df.iterrows():
    ws.append([
        row['fornitore_id'],
        row['fornitore_name'],
        row['balance_fornitore'],
        row['tipo_anomalia'],
        row['documento'],
        row['data'],
        row['importo'],
        row['residuo'],
        row['status']
    ])

# Auto-width
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

# Sheet Statistiche
ws_stats = wb.create_sheet("Statistiche")
ws_stats.append(['STATISTICHE ANOMALIE FORNITORI'])
ws_stats.append(['Data Analisi', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
ws_stats.append([])
ws_stats.append(['Totale Anomalie', len(df)])
ws_stats.append(['Valore Totale Anomalie', f'CHF {df["importo"].sum():,.2f}'])
ws_stats.append([])
ws_stats.append(['DISTRIBUZIONE PER TIPO'])
ws_stats.append(['Tipo Anomalia', 'Count', 'Valore CHF'])

for tipo, row in tipo_dist.iterrows():
    ws_stats.append([tipo, row['count'], row['importo']])

ws_stats.append([])
ws_stats.append(['TOP 10 FORNITORI PER NUMERO ANOMALIE'])
ws_stats.append(['Fornitore', 'Numero Anomalie'])

for fornitore, count in top_count.items():
    ws_stats.append([fornitore, count])

ws_stats.append([])
ws_stats.append(['TOP 10 FORNITORI PER VALORE ANOMALIE'])
ws_stats.append(['Fornitore', 'Valore CHF'])

for fornitore, value in top_value.items():
    ws_stats.append([fornitore, value])

# Sheet Piano Azione
ws_action = wb.create_sheet("Piano Azione")
ws_action.append(['PIANO AZIONE RICONCILIAZIONE FORNITORI'])
ws_action.append([])
ws_action.append(['Step', 'Azione', 'Responsabile', 'Scadenza', 'Status'])

actions = [
    ['1', 'Richiedere estratti conto TOP 10 fornitori', 'Responsabile Acquisti', '7 giorni', 'TODO'],
    ['2', 'Ricerca fatture mancanti in email/archivio', 'Contabilità', '7 giorni', 'TODO'],
    ['3', 'Registrare fatture mancanti in Odoo', 'Contabilità', '14 giorni', 'TODO'],
    ['4', 'Collegare pagamenti a fatture (TOP 10)', 'Contabilità', '14 giorni', 'TODO'],
    ['5', 'Verificare duplicati pagamenti', 'Contabilità', '14 giorni', 'TODO'],
    ['6', 'Riclassificare anticipi su conto separato', 'Contabilità', '21 giorni', 'TODO'],
    ['7', 'Completare riconciliazione tutti fornitori', 'Contabilità', '60 giorni', 'TODO'],
    ['8', 'Training team workflow Odoo', 'CFO', '30 giorni', 'TODO'],
    ['9', 'Implementare controlli automatici', 'IT + Contabilità', '60 giorni', 'TODO'],
    ['10', 'Review settimanale riconciliazioni', 'CFO', 'Ongoing', 'TODO']
]

for action in actions:
    ws_action.append(action)

# Header styling
for cell in ws_action[1]:
    cell.font = Font(bold=True, size=12)

for cell in ws_action[3]:
    cell.fill = header_fill
    cell.font = header_font

# Salva
filename = 'c:\\Users\\lapa\\Desktop\\Claude Code\\app-hub-platform\\ANOMALIE-FORNITORI-DETTAGLIO-2024.xlsx'
wb.save(filename)

print(f"\nReport salvato: {filename}")
print("\n" + "="*80)
print("ANALISI COMPLETATA")
print("="*80)
print(f"\nFile generati:")
print(f"1. {filename}")
print(f"   - Sheet 'Anomalie Fornitori': {len(df)} righe")
print(f"   - Sheet 'Statistiche': Analisi aggregate")
print(f"   - Sheet 'Piano Azione': 10 step per risoluzione")
print("\nUsare questo file per task force riconciliazione fornitori!")
