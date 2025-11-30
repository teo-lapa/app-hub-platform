#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
RICONCILIAZIONE CLIENTI E FORNITORI 2024
Analisi completa crediti/debiti con aging report e discrepanze
"""

import xmlrpc.client
import ssl
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import json
import sys
import io

# Fix encoding Windows
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# Configurazione Odoo
ODOO_URL = "https://lapadevadmin-lapa-v2-staging-2406-25408900.dev.odoo.com"
ODOO_DB = "lapadevadmin-lapa-v2-staging-2406-25408900"
ODOO_USERNAME = "paul@lapa.ch"
ODOO_PASSWORD = "lapa201180"

# Data di chiusura
DATA_CHIUSURA = "2024-12-31"

def connect_odoo():
    """Connessione a Odoo"""
    print(f"🔗 Connessione a Odoo: {ODOO_URL}")

    context = ssl._create_unverified_context()
    common = xmlrpc.client.ServerProxy(
        f'{ODOO_URL}/xmlrpc/2/common',
        context=context,
        allow_none=True
    )

    uid = common.authenticate(ODOO_DB, ODOO_USERNAME, ODOO_PASSWORD, {})

    if not uid:
        raise Exception("❌ Autenticazione fallita")

    print(f"✅ Autenticato come UID: {uid}")

    models = xmlrpc.client.ServerProxy(
        f'{ODOO_URL}/xmlrpc/2/object',
        context=context,
        allow_none=True
    )

    return uid, models

def get_account_info(models, uid, account_code):
    """Recupera informazioni account"""
    accounts = models.execute_kw(
        ODOO_DB, uid, ODOO_PASSWORD,
        'account.account', 'search_read',
        [[['code', '=', account_code]]],
        {'fields': ['id', 'code', 'name', 'account_type']}
    )

    if not accounts:
        raise Exception(f"❌ Account {account_code} non trovato")

    return accounts[0]

def get_account_moves(models, uid, account_id, date_to):
    """Recupera tutti i movimenti contabili per un account"""
    print(f"📊 Recupero movimenti account {account_id} fino a {date_to}")

    moves = models.execute_kw(
        ODOO_DB, uid, ODOO_PASSWORD,
        'account.move.line', 'search_read',
        [[
            ['account_id', '=', account_id],
            ['date', '<=', date_to],
            ['parent_state', '=', 'posted']  # Solo movimenti contabilizzati
        ]],
        {
            'fields': [
                'id', 'date', 'name', 'ref', 'partner_id',
                'debit', 'credit', 'balance', 'amount_currency',
                'currency_id', 'move_id', 'move_name',
                'matched_debit_ids', 'matched_credit_ids',
                'reconciled', 'full_reconcile_id'
            ],
            'order': 'date asc, id asc'
        }
    )

    print(f"✅ Trovati {len(moves)} movimenti")
    return moves

def get_partner_balances(models, uid, account_id, date_to, partner_type):
    """Calcola saldi per partner con aging"""
    print(f"📊 Calcolo saldi {partner_type} al {date_to}")

    moves = get_account_moves(models, uid, account_id, date_to)

    # Raggruppa per partner
    partner_data = {}

    for move in moves:
        partner_id = move.get('partner_id')
        if not partner_id:
            partner_id = [0, 'SENZA PARTNER']

        pid = partner_id[0]
        pname = partner_id[1] if isinstance(partner_id, list) else 'Unknown'

        if pid not in partner_data:
            partner_data[pid] = {
                'id': pid,
                'name': pname,
                'balance': 0,
                'debit': 0,
                'credit': 0,
                'moves': [],
                'open_moves': []
            }

        balance = move.get('balance', 0) or 0
        debit = move.get('debit', 0) or 0
        credit = move.get('credit', 0) or 0

        partner_data[pid]['balance'] += balance
        partner_data[pid]['debit'] += debit
        partner_data[pid]['credit'] += credit
        partner_data[pid]['moves'].append(move)

        # Se non riconciliato, è aperto
        if not move.get('reconciled', False):
            partner_data[pid]['open_moves'].append(move)

    # Filtra solo partner con saldo != 0
    result = []
    for pid, data in partner_data.items():
        if abs(data['balance']) > 0.01:  # Tolleranza centesimi
            # Calcola aging
            aging = calculate_aging(data['open_moves'], date_to)

            result.append({
                'partner_id': data['id'],
                'partner_name': data['name'],
                'balance': round(data['balance'], 2),
                'debit': round(data['debit'], 2),
                'credit': round(data['credit'], 2),
                'total_moves': len(data['moves']),
                'open_moves': len(data['open_moves']),
                'current': round(aging['current'], 2),
                '30_days': round(aging['30_days'], 2),
                '60_days': round(aging['60_days'], 2),
                '90_days': round(aging['90_days'], 2),
                'over_90': round(aging['over_90'], 2)
            })

    # Ordina per saldo decrescente (valore assoluto)
    result.sort(key=lambda x: abs(x['balance']), reverse=True)

    print(f"✅ Trovati {len(result)} {partner_type} con saldo aperto")
    return result

def calculate_aging(open_moves, date_to):
    """Calcola aging dei movimenti aperti"""
    aging = {
        'current': 0,
        '30_days': 0,
        '60_days': 0,
        '90_days': 0,
        'over_90': 0
    }

    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')

    for move in open_moves:
        move_date = datetime.strptime(move['date'], '%Y-%m-%d')
        days_old = (date_to_obj - move_date).days
        balance = move.get('balance', 0) or 0

        if days_old <= 30:
            aging['current'] += balance
        elif days_old <= 60:
            aging['30_days'] += balance
        elif days_old <= 90:
            aging['60_days'] += balance
        elif days_old <= 120:
            aging['90_days'] += balance
        else:
            aging['over_90'] += balance

    return aging

def get_invoices_vs_payments(models, uid, partner_id, partner_type, year=2024):
    """Analizza fatture vs pagamenti per un partner"""
    print(f"🔍 Analisi {partner_type} ID {partner_id} - Anno {year}")

    if partner_type == 'cliente':
        # Fatture clienti (out_invoice)
        invoice_type = ['out_invoice', 'out_refund']
        payment_type = 'inbound'
    else:
        # Fatture fornitori (in_invoice)
        invoice_type = ['in_invoice', 'in_refund']
        payment_type = 'outbound'

    # Recupera fatture
    invoices = models.execute_kw(
        ODOO_DB, uid, ODOO_PASSWORD,
        'account.move', 'search_read',
        [[
            ['partner_id', '=', partner_id],
            ['move_type', 'in', invoice_type],
            ['state', '=', 'posted'],
            ['invoice_date', '>=', f'{year}-01-01'],
            ['invoice_date', '<=', f'{year}-12-31']
        ]],
        {
            'fields': [
                'id', 'name', 'invoice_date', 'invoice_date_due',
                'move_type', 'amount_total', 'amount_residual',
                'payment_state', 'state'
            ]
        }
    )

    # Recupera pagamenti
    payments = models.execute_kw(
        ODOO_DB, uid, ODOO_PASSWORD,
        'account.payment', 'search_read',
        [[
            ['partner_id', '=', partner_id],
            ['payment_type', '=', payment_type],
            ['state', '=', 'posted'],
            ['date', '>=', f'{year}-01-01'],
            ['date', '<=', f'{year}-12-31']
        ]],
        {
            'fields': [
                'id', 'name', 'date', 'amount', 'currency_id',
                'payment_type', 'partner_type', 'ref',
                'reconciled_invoice_ids'
            ]
        }
    )

    # Analizza discrepanze
    discrepanze = []

    # Fatture non incassate/pagate
    for inv in invoices:
        if abs(inv.get('amount_residual', 0) or 0) > 0.01:
            discrepanze.append({
                'type': 'Fattura non incassata' if partner_type == 'cliente' else 'Fattura non pagata',
                'document': inv['name'],
                'date': inv['invoice_date'],
                'amount': inv['amount_total'],
                'residual': inv['amount_residual'],
                'status': inv['payment_state']
            })

    # Pagamenti senza fattura collegata
    for pay in payments:
        reconciled = pay.get('reconciled_invoice_ids', [])
        if not reconciled or len(reconciled) == 0:
            discrepanze.append({
                'type': 'Pagamento senza fattura',
                'document': pay['name'],
                'date': pay['date'],
                'amount': pay['amount'],
                'residual': 0,
                'status': 'not_linked'
            })

    return {
        'invoices': invoices,
        'payments': payments,
        'discrepanze': discrepanze,
        'total_invoiced': sum(i.get('amount_total', 0) or 0 for i in invoices),
        'total_paid': sum(p.get('amount', 0) or 0 for p in payments),
        'total_residual': sum(i.get('amount_residual', 0) or 0 for i in invoices)
    }

def create_excel_report(clienti_data, fornitori_data, top_clienti_analysis, top_fornitori_analysis):
    """Crea report Excel completo"""
    print("📊 Generazione report Excel...")

    wb = Workbook()
    wb.remove(wb.active)

    # Stili
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 1. SHEET CLIENTI
    ws_clienti = wb.create_sheet("Clienti - Saldi")
    ws_clienti.append([
        'Partner ID', 'Cliente', 'Saldo', 'Dare', 'Avere',
        'Tot.Movimenti', 'Mov.Aperti', 'Corrente', '30gg', '60gg', '90gg', '>90gg'
    ])

    # Header styling
    for cell in ws_clienti[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Dati clienti
    total_clienti = 0
    for row in clienti_data:
        ws_clienti.append([
            row['partner_id'],
            row['partner_name'],
            row['balance'],
            row['debit'],
            row['credit'],
            row['total_moves'],
            row['open_moves'],
            row['current'],
            row['30_days'],
            row['60_days'],
            row['90_days'],
            row['over_90']
        ])
        total_clienti += row['balance']

    # Totale
    last_row = ws_clienti.max_row + 1
    ws_clienti.append(['', 'TOTALE', total_clienti, '', '', '', '', '', '', '', '', ''])
    for cell in ws_clienti[last_row]:
        cell.font = Font(bold=True)
        cell.fill = success_fill if total_clienti > 0 else warning_fill

    # Auto-width
    for column in ws_clienti.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_clienti.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # 2. SHEET FORNITORI
    ws_fornitori = wb.create_sheet("Fornitori - Saldi")
    ws_fornitori.append([
        'Partner ID', 'Fornitore', 'Saldo', 'Dare', 'Avere',
        'Tot.Movimenti', 'Mov.Aperti', 'Corrente', '30gg', '60gg', '90gg', '>90gg'
    ])

    for cell in ws_fornitori[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    total_fornitori = 0
    for row in fornitori_data:
        ws_fornitori.append([
            row['partner_id'],
            row['partner_name'],
            row['balance'],
            row['debit'],
            row['credit'],
            row['total_moves'],
            row['open_moves'],
            row['current'],
            row['30_days'],
            row['60_days'],
            row['90_days'],
            row['over_90']
        ])
        total_fornitori += row['balance']

    last_row = ws_fornitori.max_row + 1
    ws_fornitori.append(['', 'TOTALE', total_fornitori, '', '', '', '', '', '', '', '', ''])
    for cell in ws_fornitori[last_row]:
        cell.font = Font(bold=True)
        cell.fill = warning_fill if total_fornitori < 0 else success_fill

    for column in ws_fornitori.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_fornitori.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # 3. SHEET DISCREPANZE CLIENTI
    ws_disc_clienti = wb.create_sheet("Discrepanze Clienti")
    ws_disc_clienti.append(['Cliente', 'Tipo Anomalia', 'Documento', 'Data', 'Importo', 'Residuo', 'Status'])

    for cell in ws_disc_clienti[1]:
        cell.fill = header_fill
        cell.font = header_font

    for cliente_analysis in top_clienti_analysis:
        cliente_name = cliente_analysis['partner_name']
        for disc in cliente_analysis['analysis']['discrepanze']:
            ws_disc_clienti.append([
                cliente_name,
                disc['type'],
                disc['document'],
                disc['date'],
                disc['amount'],
                disc['residual'],
                disc['status']
            ])

    # 4. SHEET DISCREPANZE FORNITORI
    ws_disc_fornitori = wb.create_sheet("Discrepanze Fornitori")
    ws_disc_fornitori.append(['Fornitore', 'Tipo Anomalia', 'Documento', 'Data', 'Importo', 'Residuo', 'Status'])

    for cell in ws_disc_fornitori[1]:
        cell.fill = header_fill
        cell.font = header_font

    for fornitore_analysis in top_fornitori_analysis:
        fornitore_name = fornitore_analysis['partner_name']
        for disc in fornitore_analysis['analysis']['discrepanze']:
            ws_disc_fornitori.append([
                fornitore_name,
                disc['type'],
                disc['document'],
                disc['date'],
                disc['amount'],
                disc['residual'],
                disc['status']
            ])

    # 5. SHEET AZIONI
    ws_azioni = wb.create_sheet("Azioni Correttive")
    ws_azioni.append(['CATEGORIA', 'AZIONE', 'PRIORITÀ', 'RESPONSABILE', 'SCADENZA'])

    for cell in ws_azioni[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Azioni clienti
    clienti_scaduti = sum(1 for c in clienti_data if c['over_90'] > 0)
    if clienti_scaduti > 0:
        ws_azioni.append([
            'CLIENTI',
            f'Sollecitare {clienti_scaduti} clienti con crediti >90gg',
            'ALTA',
            'Responsabile crediti',
            'Entro 7 giorni'
        ])

    # Azioni fornitori
    fornitori_scaduti = sum(1 for f in fornitori_data if f['over_90'] < 0)
    if fornitori_scaduti > 0:
        ws_azioni.append([
            'FORNITORI',
            f'Pagare {fornitori_scaduti} fornitori con debiti >90gg',
            'ALTA',
            'Responsabile pagamenti',
            'Entro 14 giorni'
        ])

    # Azioni discrepanze
    total_disc_clienti = sum(len(c['analysis']['discrepanze']) for c in top_clienti_analysis)
    if total_disc_clienti > 0:
        ws_azioni.append([
            'CLIENTI',
            f'Riconciliare {total_disc_clienti} discrepanze clienti',
            'MEDIA',
            'Contabilità',
            'Entro 30 giorni'
        ])

    total_disc_fornitori = sum(len(f['analysis']['discrepanze']) for f in top_fornitori_analysis)
    if total_disc_fornitori > 0:
        ws_azioni.append([
            'FORNITORI',
            f'Riconciliare {total_disc_fornitori} discrepanze fornitori',
            'MEDIA',
            'Contabilità',
            'Entro 30 giorni'
        ])

    # 6. SHEET SUMMARY
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary.append(['RICONCILIAZIONE CLIENTI E FORNITORI 2024'])
    ws_summary.append(['Data Report', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    ws_summary.append(['Data Chiusura', DATA_CHIUSURA])
    ws_summary.append([])

    ws_summary.append(['CLIENTI (Conto 1100 - Crediti)'])
    ws_summary.append(['Totale Clienti con saldo aperto', len(clienti_data)])
    ws_summary.append(['Totale Crediti', f'CHF {total_clienti:,.2f}'])
    ws_summary.append(['Crediti >90gg', f'CHF {sum(c["over_90"] for c in clienti_data):,.2f}'])
    ws_summary.append([])

    ws_summary.append(['FORNITORI (Conto 2000 - Debiti)'])
    ws_summary.append(['Totale Fornitori con saldo aperto', len(fornitori_data)])
    ws_summary.append(['Totale Debiti', f'CHF {abs(total_fornitori):,.2f}'])
    ws_summary.append(['Debiti >90gg', f'CHF {abs(sum(f["over_90"] for f in fornitori_data)):,.2f}'])
    ws_summary.append([])

    ws_summary.append(['DISCREPANZE'])
    ws_summary.append(['Discrepanze Clienti', total_disc_clienti])
    ws_summary.append(['Discrepanze Fornitori', total_disc_fornitori])
    ws_summary.append([])

    ws_summary.append(['AZIONI RICHIESTE'])
    ws_summary.append(['Solleciti clienti scaduti', clienti_scaduti])
    ws_summary.append(['Pagamenti fornitori scaduti', fornitori_scaduti])

    # Styling summary
    ws_summary['A1'].font = Font(bold=True, size=14)
    for row in range(5, 20):
        cell = ws_summary[f'A{row}']
        if cell.value and isinstance(cell.value, str) and cell.value.isupper():
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Salva
    filename = f"c:\\Users\\lapa\\Desktop\\Claude Code\\app-hub-platform\\RICONCILIAZIONE-CLIENTI-FORNITORI-2024.xlsx"
    wb.save(filename)
    print(f"✅ Report salvato: {filename}")
    return filename

def main():
    """Esecuzione principale"""
    print("=" * 80)
    print("RICONCILIAZIONE CLIENTI E FORNITORI 2024")
    print("=" * 80)

    # Connessione
    uid, models = connect_odoo()

    # Recupera info accounts
    print("\n📋 STEP 1: Recupero informazioni accounts")
    account_clienti = get_account_info(models, uid, '1100')
    print(f"✅ Account Clienti: {account_clienti['code']} - {account_clienti['name']}")

    account_fornitori = get_account_info(models, uid, '2000')
    print(f"✅ Account Fornitori: {account_fornitori['code']} - {account_fornitori['name']}")

    # Analisi clienti
    print("\n📊 STEP 2: Analisi Crediti Clienti")
    clienti_data = get_partner_balances(
        models, uid, account_clienti['id'], DATA_CHIUSURA, 'clienti'
    )

    total_crediti = sum(c['balance'] for c in clienti_data)
    print(f"💰 Totale Crediti: CHF {total_crediti:,.2f}")
    print(f"👥 Clienti con saldo aperto: {len(clienti_data)}")

    # Analisi fornitori
    print("\n📊 STEP 3: Analisi Debiti Fornitori")
    fornitori_data = get_partner_balances(
        models, uid, account_fornitori['id'], DATA_CHIUSURA, 'fornitori'
    )

    total_debiti = sum(f['balance'] for f in fornitori_data)
    print(f"💰 Totale Debiti: CHF {abs(total_debiti):,.2f}")
    print(f"👥 Fornitori con saldo aperto: {len(fornitori_data)}")

    # Analisi dettagliata TOP 20 clienti
    print("\n🔍 STEP 4: Analisi dettagliata TOP 20 Clienti")
    top_clienti_analysis = []
    for i, cliente in enumerate(clienti_data[:20], 1):
        print(f"  [{i}/20] {cliente['partner_name']} - CHF {cliente['balance']:,.2f}")
        analysis = get_invoices_vs_payments(
            models, uid, cliente['partner_id'], 'cliente', 2024
        )
        top_clienti_analysis.append({
            'partner_id': cliente['partner_id'],
            'partner_name': cliente['partner_name'],
            'balance': cliente['balance'],
            'analysis': analysis
        })

    # Analisi dettagliata TOP 20 fornitori
    print("\n🔍 STEP 5: Analisi dettagliata TOP 20 Fornitori")
    top_fornitori_analysis = []
    for i, fornitore in enumerate(fornitori_data[:20], 1):
        print(f"  [{i}/20] {fornitore['partner_name']} - CHF {abs(fornitore['balance']):,.2f}")
        analysis = get_invoices_vs_payments(
            models, uid, fornitore['partner_id'], 'fornitore', 2024
        )
        top_fornitori_analysis.append({
            'partner_id': fornitore['partner_id'],
            'partner_name': fornitore['partner_name'],
            'balance': fornitore['balance'],
            'analysis': analysis
        })

    # Genera report Excel
    print("\n📊 STEP 6: Generazione Report Excel")
    filename = create_excel_report(
        clienti_data, fornitori_data,
        top_clienti_analysis, top_fornitori_analysis
    )

    # Summary finale
    print("\n" + "=" * 80)
    print("✅ RICONCILIAZIONE COMPLETATA")
    print("=" * 80)
    print(f"\n📊 CLIENTI (Conto 1100):")
    print(f"   - Totale clienti con saldo: {len(clienti_data)}")
    print(f"   - Totale crediti: CHF {total_crediti:,.2f}")
    print(f"   - Crediti >90gg: CHF {sum(c['over_90'] for c in clienti_data):,.2f}")

    print(f"\n📊 FORNITORI (Conto 2000):")
    print(f"   - Totale fornitori con saldo: {len(fornitori_data)}")
    print(f"   - Totale debiti: CHF {abs(total_debiti):,.2f}")
    print(f"   - Debiti >90gg: CHF {abs(sum(f['over_90'] for f in fornitori_data)):,.2f}")

    total_disc_clienti = sum(len(c['analysis']['discrepanze']) for c in top_clienti_analysis)
    total_disc_fornitori = sum(len(f['analysis']['discrepanze']) for f in top_fornitori_analysis)

    print(f"\n⚠️ DISCREPANZE:")
    print(f"   - Discrepanze clienti (TOP 20): {total_disc_clienti}")
    print(f"   - Discrepanze fornitori (TOP 20): {total_disc_fornitori}")

    print(f"\n📄 Report Excel: {filename}")
    print("\n" + "=" * 80)

    # Salva anche JSON per analisi ulteriori
    json_data = {
        'data_report': datetime.now().isoformat(),
        'data_chiusura': DATA_CHIUSURA,
        'clienti': {
            'count': len(clienti_data),
            'total_crediti': total_crediti,
            'data': clienti_data
        },
        'fornitori': {
            'count': len(fornitori_data),
            'total_debiti': total_debiti,
            'data': fornitori_data
        },
        'top_clienti_analysis': top_clienti_analysis,
        'top_fornitori_analysis': top_fornitori_analysis
    }

    json_filename = "c:\\Users\\lapa\\Desktop\\Claude Code\\app-hub-platform\\riconciliazione-clienti-fornitori-2024.json"
    with open(json_filename, 'w', encoding='utf-8') as f:
        json.dump(json_data, f, indent=2, ensure_ascii=False, default=str)
    print(f"💾 Dati JSON salvati: {json_filename}")

if __name__ == "__main__":
    main()
