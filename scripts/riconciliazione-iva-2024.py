#!/usr/bin/env python3
"""
RICONCILIAZIONE IVA 2024 COMPLETA
Analisi quadratura IVA vendite/acquisti e identificazione errori
"""

import xmlrpc.client
from datetime import datetime, date
from collections import defaultdict
import json
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configurazione Odoo
ODOO_URL = 'https://lapadevadmin-lapa-v2-staging-2406-25408900.dev.odoo.com'
ODOO_DB = 'lapadevadmin-lapa-v2-staging-2406-25408900'
ODOO_USERNAME = 'paul@lapa.ch'
ODOO_PASSWORD = 'lapa201180'

# Aliquote IVA standard Svizzera
IVA_ALIQUOTE = {
    7.7: 'IVA normale 7.7%',
    2.5: 'IVA ridotta 2.5%',
    3.7: 'IVA alloggio 3.7%',
    0.0: 'Esente/Fuori campo'
}

class OdooVATReconciliation:
    def __init__(self):
        print("Connessione a Odoo...")
        self.common = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/common')
        self.uid = self.common.authenticate(ODOO_DB, ODOO_USERNAME, ODOO_PASSWORD, {})

        if not self.uid:
            raise Exception("Autenticazione Odoo fallita!")

        self.models = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/object')
        print(f"Connesso come user ID: {self.uid}")

        self.vat_data = {
            'vendite_mensili': defaultdict(lambda: defaultdict(Decimal)),
            'acquisti_mensili': defaultdict(lambda: defaultdict(Decimal)),
            'errori': [],
            'dettagli_vendite': [],
            'dettagli_acquisti': []
        }

    def execute(self, model, method, domain=None, fields=None, **kwargs):
        """Wrapper per chiamate Odoo"""
        if domain is None:
            domain = []
        args = [domain]
        if fields:
            kwargs['fields'] = fields
        return self.models.execute_kw(
            ODOO_DB, self.uid, ODOO_PASSWORD,
            model, method, args, kwargs
        )

    def get_vat_amount(self, amount, rate):
        """Calcola IVA da importo totale"""
        if rate == 0:
            return Decimal('0')
        # IVA = Totale * (rate / (100 + rate))
        return Decimal(str(amount)) * Decimal(str(rate)) / (Decimal('100') + Decimal(str(rate)))

    def analyze_sales_vat(self):
        """1. ANALISI IVA A DEBITO (vendite) - Conti 2200-2299"""
        print("\n" + "="*80)
        print("1. ANALISI IVA A DEBITO (VENDITE)")
        print("="*80)

        # Trova tutti i conti IVA vendite
        vat_accounts = self.execute(
            'account.account', 'search_read',
            domain=[['code', '>=', '2200'], ['code', '<', '2300']],
            fields=['id', 'code', 'name']
        )

        print(f"\nConti IVA vendite trovati: {len(vat_accounts)}")
        for acc in vat_accounts:
            print(f"  - {acc['code']}: {acc['name']}")

        account_ids = [acc['id'] for acc in vat_accounts]

        # Estrai movimenti contabili 2024
        domain = [
            ['account_id', 'in', account_ids],
            ['date', '>=', '2024-01-01'],
            ['date', '<=', '2024-12-31'],
            ['parent_state', '=', 'posted']  # Solo movimenti confermati
        ]

        moves = self.execute(
            'account.move.line', 'search_read',
            domain=domain,
            fields=['id', 'date', 'name', 'debit', 'credit', 'account_id',
                   'move_id', 'partner_id', 'tax_ids', 'tax_base_amount']
        )

        print(f"\nMovimenti contabili trovati: {len(moves)}")

        # Analizza per mese e aliquota
        for move in moves:
            date_obj = datetime.strptime(move['date'], '%Y-%m-%d')
            mese = date_obj.strftime('%Y-%m')

            # IVA a debito = credit (non debit!)
            iva_amount = Decimal(str(move['credit'])) - Decimal(str(move['debit']))

            # Determina aliquota (dalla base imponibile se disponibile)
            aliquota = 'Non classificata'
            if move.get('tax_base_amount') and move['tax_base_amount'] != 0:
                # Calcola aliquota effettiva
                base = Decimal(str(move['tax_base_amount']))
                if base != 0:
                    rate = (iva_amount / base) * 100
                    # Arrotonda alle aliquote standard
                    if abs(rate - Decimal('7.7')) < Decimal('0.1'):
                        aliquota = '7.7%'
                    elif abs(rate - Decimal('2.5')) < Decimal('0.1'):
                        aliquota = '2.5%'
                    elif abs(rate - Decimal('3.7')) < Decimal('0.1'):
                        aliquota = '3.7%'
                    elif abs(rate) < Decimal('0.1'):
                        aliquota = '0.0%'
                    else:
                        aliquota = f'{float(rate):.2f}%'
                        # ERRORE: Aliquota non standard!
                        self.vat_data['errori'].append({
                            'tipo': 'ALIQUOTA_NON_STANDARD',
                            'data': move['date'],
                            'descrizione': move['name'],
                            'aliquota_calcolata': f'{float(rate):.2f}%',
                            'base_imponibile': float(base),
                            'iva': float(iva_amount),
                            'move_id': move['move_id'][0] if move['move_id'] else None
                        })

            self.vat_data['vendite_mensili'][mese][aliquota] += iva_amount

            # Salva dettagli
            self.vat_data['dettagli_vendite'].append({
                'data': move['date'],
                'mese': mese,
                'descrizione': move['name'],
                'partner': move['partner_id'][1] if move['partner_id'] else '',
                'conto': move['account_id'][1],
                'base_imponibile': float(move.get('tax_base_amount', 0)),
                'iva': float(iva_amount),
                'aliquota': aliquota,
                'move_id': move['move_id'][0] if move['move_id'] else None
            })

        # Stampa riepilogo vendite
        print("\nRIEPILOGO IVA VENDITE PER MESE:")
        total_vendite = Decimal('0')
        for mese in sorted(self.vat_data['vendite_mensili'].keys()):
            totale_mese = sum(self.vat_data['vendite_mensili'][mese].values())
            total_vendite += totale_mese
            print(f"\n{mese}: CHF {totale_mese:,.2f}")
            for aliquota, importo in sorted(self.vat_data['vendite_mensili'][mese].items()):
                print(f"  {aliquota}: CHF {importo:,.2f}")

        print(f"\nTOTALE IVA VENDITE 2024: CHF {total_vendite:,.2f}")
        return total_vendite

    def analyze_purchases_vat(self):
        """2. ANALISI IVA A CREDITO (acquisti) - Conti 1170-1179"""
        print("\n" + "="*80)
        print("2. ANALISI IVA A CREDITO (ACQUISTI)")
        print("="*80)

        # Trova tutti i conti IVA acquisti
        vat_accounts = self.execute(
            'account.account', 'search_read',
            domain=[['code', '>=', '1170'], ['code', '<', '1180']],
            fields=['id', 'code', 'name']
        )

        print(f"\nConti IVA acquisti trovati: {len(vat_accounts)}")
        for acc in vat_accounts:
            print(f"  - {acc['code']}: {acc['name']}")

        account_ids = [acc['id'] for acc in vat_accounts]

        # Estrai movimenti contabili 2024
        domain = [
            ['account_id', 'in', account_ids],
            ['date', '>=', '2024-01-01'],
            ['date', '<=', '2024-12-31'],
            ['parent_state', '=', 'posted']
        ]

        moves = self.execute(
            'account.move.line', 'search_read',
            domain=domain,
            fields=['id', 'date', 'name', 'debit', 'credit', 'account_id',
                   'move_id', 'partner_id', 'tax_ids', 'tax_base_amount']
        )

        print(f"\nMovimenti contabili trovati: {len(moves)}")

        # Analizza per mese e aliquota
        for move in moves:
            date_obj = datetime.strptime(move['date'], '%Y-%m-%d')
            mese = date_obj.strftime('%Y-%m')

            # IVA a credito = debit (non credit!)
            iva_amount = Decimal(str(move['debit'])) - Decimal(str(move['credit']))

            # Determina aliquota
            aliquota = 'Non classificata'
            if move.get('tax_base_amount') and move['tax_base_amount'] != 0:
                base = Decimal(str(move['tax_base_amount']))
                if base != 0:
                    rate = (iva_amount / base) * 100
                    if abs(rate - Decimal('7.7')) < Decimal('0.1'):
                        aliquota = '7.7%'
                    elif abs(rate - Decimal('2.5')) < Decimal('0.1'):
                        aliquota = '2.5%'
                    elif abs(rate - Decimal('3.7')) < Decimal('0.1'):
                        aliquota = '3.7%'
                    elif abs(rate) < Decimal('0.1'):
                        aliquota = '0.0%'
                    else:
                        aliquota = f'{float(rate):.2f}%'
                        # ERRORE: Aliquota non standard!
                        self.vat_data['errori'].append({
                            'tipo': 'ALIQUOTA_NON_STANDARD_ACQUISTI',
                            'data': move['date'],
                            'descrizione': move['name'],
                            'aliquota_calcolata': f'{float(rate):.2f}%',
                            'base_imponibile': float(base),
                            'iva': float(iva_amount),
                            'move_id': move['move_id'][0] if move['move_id'] else None
                        })

            self.vat_data['acquisti_mensili'][mese][aliquota] += iva_amount

            # Salva dettagli
            self.vat_data['dettagli_acquisti'].append({
                'data': move['date'],
                'mese': mese,
                'descrizione': move['name'],
                'partner': move['partner_id'][1] if move['partner_id'] else '',
                'conto': move['account_id'][1],
                'base_imponibile': float(move.get('tax_base_amount', 0)),
                'iva': float(iva_amount),
                'aliquota': aliquota,
                'move_id': move['move_id'][0] if move['move_id'] else None
            })

        # Stampa riepilogo acquisti
        print("\nRIEPILOGO IVA ACQUISTI PER MESE:")
        total_acquisti = Decimal('0')
        for mese in sorted(self.vat_data['acquisti_mensili'].keys()):
            totale_mese = sum(self.vat_data['acquisti_mensili'][mese].values())
            total_acquisti += totale_mese
            print(f"\n{mese}: CHF {totale_mese:,.2f}")
            for aliquota, importo in sorted(self.vat_data['acquisti_mensili'][mese].items()):
                print(f"  {aliquota}: CHF {importo:,.2f}")

        print(f"\nTOTALE IVA ACQUISTI 2024: CHF {total_acquisti:,.2f}")
        return total_acquisti

    def find_vat_errors(self):
        """4. IDENTIFICAZIONE ERRORI IVA"""
        print("\n" + "="*80)
        print("4. IDENTIFICAZIONE ERRORI IVA")
        print("="*80)

        # Cerca fatture duplicate (stesso importo IVA nello stesso giorno)
        vendite_by_date = defaultdict(list)
        for det in self.vat_data['dettagli_vendite']:
            vendite_by_date[det['data']].append(det)

        for data, fatture in vendite_by_date.items():
            if len(fatture) > 1:
                # Controlla importi identici
                importi = defaultdict(list)
                for fatt in fatture:
                    importi[fatt['iva']].append(fatt)

                for iva, duplicati in importi.items():
                    if len(duplicati) > 1 and iva > 0:
                        self.vat_data['errori'].append({
                            'tipo': 'POSSIBILE_DUPLICATO',
                            'data': data,
                            'descrizione': f"{len(duplicati)} fatture con IVA identica: CHF {iva:.2f}",
                            'fatture': [d['descrizione'] for d in duplicati],
                            'move_ids': [d['move_id'] for d in duplicati]
                        })

        # Cerca IVA su operazioni che dovrebbero essere esenti
        for det in self.vat_data['dettagli_vendite']:
            if det['descrizione'] and isinstance(det['descrizione'], str):
                desc_lower = det['descrizione'].lower()
                # Keywords che indicano operazioni esenti
                esenti_keywords = ['export', 'estero', 'intrastat', 'reverse charge']
                if any(kw in desc_lower for kw in esenti_keywords):
                    if det['iva'] != 0:
                        self.vat_data['errori'].append({
                            'tipo': 'IVA_SU_OPERAZIONE_ESENTE',
                            'data': det['data'],
                            'descrizione': det['descrizione'],
                            'iva_applicata': det['iva'],
                            'move_id': det['move_id']
                        })

        # Cerca base imponibile = 0 ma IVA > 0 (errore!)
        for det in self.vat_data['dettagli_vendite'] + self.vat_data['dettagli_acquisti']:
            if det['base_imponibile'] == 0 and det['iva'] != 0:
                self.vat_data['errori'].append({
                    'tipo': 'IVA_SENZA_BASE_IMPONIBILE',
                    'data': det['data'],
                    'descrizione': det['descrizione'],
                    'iva': det['iva'],
                    'move_id': det['move_id']
                })

        print(f"\nTOTALE ERRORI IDENTIFICATI: {len(self.vat_data['errori'])}")

        # Raggruppa errori per tipo
        errori_per_tipo = defaultdict(int)
        for err in self.vat_data['errori']:
            errori_per_tipo[err['tipo']] += 1

        for tipo, count in sorted(errori_per_tipo.items()):
            print(f"  - {tipo}: {count}")

    def calculate_quarterly_summary(self):
        """3. QUADRATURA TRIMESTRALE"""
        print("\n" + "="*80)
        print("3. QUADRATURA IVA TRIMESTRALE")
        print("="*80)

        trimestri = {
            'Q1 2024': ['2024-01', '2024-02', '2024-03'],
            'Q2 2024': ['2024-04', '2024-05', '2024-06'],
            'Q3 2024': ['2024-07', '2024-08', '2024-09'],
            'Q4 2024': ['2024-10', '2024-11', '2024-12']
        }

        quadratura = {}

        for trimestre, mesi in trimestri.items():
            vendite = Decimal('0')
            acquisti = Decimal('0')

            for mese in mesi:
                vendite += sum(self.vat_data['vendite_mensili'].get(mese, {}).values())
                acquisti += sum(self.vat_data['acquisti_mensili'].get(mese, {}).values())

            saldo = vendite - acquisti

            quadratura[trimestre] = {
                'vendite': vendite,
                'acquisti': acquisti,
                'saldo': saldo,
                'stato': 'A DEBITO' if saldo > 0 else 'A CREDITO'
            }

            print(f"\n{trimestre}:")
            print(f"  IVA Vendite:  CHF {vendite:>12,.2f}")
            print(f"  IVA Acquisti: CHF {acquisti:>12,.2f}")
            print(f"  {'-'*35}")
            print(f"  Saldo:        CHF {saldo:>12,.2f} ({quadratura[trimestre]['stato']})")

        return quadratura

    def generate_excel_report(self, total_vendite, total_acquisti, quadratura):
        """Genera report Excel completo"""
        print("\n" + "="*80)
        print("GENERAZIONE REPORT EXCEL")
        print("="*80)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Rimuovi foglio default

        # Stili
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        error_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        warning_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 1. RIEPILOGO GENERALE
        ws1 = wb.create_sheet('Riepilogo Generale')
        ws1['A1'] = 'RICONCILIAZIONE IVA 2024 - RIEPILOGO'
        ws1['A1'].font = Font(size=16, bold=True)

        row = 3
        ws1[f'A{row}'] = 'TOTALI ANNUALI'
        ws1[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        ws1[f'A{row}'] = 'IVA Vendite (a debito)'
        ws1[f'B{row}'] = float(total_vendite)
        ws1[f'B{row}'].number_format = '#,##0.00'
        row += 1

        ws1[f'A{row}'] = 'IVA Acquisti (a credito)'
        ws1[f'B{row}'] = float(total_acquisti)
        ws1[f'B{row}'].number_format = '#,##0.00'
        row += 1

        ws1[f'A{row}'] = 'SALDO IVA 2024'
        ws1[f'B{row}'] = float(total_vendite - total_acquisti)
        ws1[f'B{row}'].number_format = '#,##0.00'
        ws1[f'A{row}'].font = Font(bold=True)
        ws1[f'B{row}'].font = Font(bold=True)
        row += 2

        ws1[f'A{row}'] = 'ERRORI IDENTIFICATI'
        ws1[f'B{row}'] = len(self.vat_data['errori'])
        ws1[f'A{row}'].font = Font(bold=True)
        ws1[f'B{row}'].font = Font(bold=True, color='FF0000')

        # 2. IVA VENDITE MENSILE
        ws2 = wb.create_sheet('IVA Vendite Mensile')
        headers = ['Mese', '7.7%', '2.5%', '3.7%', '0.0%', 'Altre', 'TOTALE']
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font

        row = 2
        for mese in sorted(self.vat_data['vendite_mensili'].keys()):
            ws2.cell(row, 1, mese)
            ws2.cell(row, 2, float(self.vat_data['vendite_mensili'][mese].get('7.7%', 0)))
            ws2.cell(row, 3, float(self.vat_data['vendite_mensili'][mese].get('2.5%', 0)))
            ws2.cell(row, 4, float(self.vat_data['vendite_mensili'][mese].get('3.7%', 0)))
            ws2.cell(row, 5, float(self.vat_data['vendite_mensili'][mese].get('0.0%', 0)))

            # Somma altre aliquote
            altre = sum(v for k, v in self.vat_data['vendite_mensili'][mese].items()
                       if k not in ['7.7%', '2.5%', '3.7%', '0.0%'])
            ws2.cell(row, 6, float(altre))

            # Totale
            totale = sum(self.vat_data['vendite_mensili'][mese].values())
            ws2.cell(row, 7, float(totale))

            # Formattazione
            for col in range(2, 8):
                ws2.cell(row, col).number_format = '#,##0.00'

            row += 1

        # 3. IVA ACQUISTI MENSILE
        ws3 = wb.create_sheet('IVA Acquisti Mensile')
        for col, header in enumerate(headers, 1):
            cell = ws3.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font

        row = 2
        for mese in sorted(self.vat_data['acquisti_mensili'].keys()):
            ws3.cell(row, 1, mese)
            ws3.cell(row, 2, float(self.vat_data['acquisti_mensili'][mese].get('7.7%', 0)))
            ws3.cell(row, 3, float(self.vat_data['acquisti_mensili'][mese].get('2.5%', 0)))
            ws3.cell(row, 4, float(self.vat_data['acquisti_mensili'][mese].get('3.7%', 0)))
            ws3.cell(row, 5, float(self.vat_data['acquisti_mensili'][mese].get('0.0%', 0)))

            altre = sum(v for k, v in self.vat_data['acquisti_mensili'][mese].items()
                       if k not in ['7.7%', '2.5%', '3.7%', '0.0%'])
            ws3.cell(row, 6, float(altre))

            totale = sum(self.vat_data['acquisti_mensili'][mese].values())
            ws3.cell(row, 7, float(totale))

            for col in range(2, 8):
                ws3.cell(row, col).number_format = '#,##0.00'

            row += 1

        # 4. QUADRATURA TRIMESTRALE
        ws4 = wb.create_sheet('Quadratura Trimestrale')
        headers = ['Trimestre', 'IVA Vendite', 'IVA Acquisti', 'Saldo', 'Stato']
        for col, header in enumerate(headers, 1):
            cell = ws4.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font

        row = 2
        for trimestre, dati in sorted(quadratura.items()):
            ws4.cell(row, 1, trimestre)
            ws4.cell(row, 2, float(dati['vendite']))
            ws4.cell(row, 3, float(dati['acquisti']))
            ws4.cell(row, 4, float(dati['saldo']))
            ws4.cell(row, 5, dati['stato'])

            for col in range(2, 5):
                ws4.cell(row, col).number_format = '#,##0.00'

            row += 1

        # 5. ERRORI IVA
        ws5 = wb.create_sheet('ERRORI IVA')
        headers = ['Tipo Errore', 'Data', 'Descrizione', 'Dettagli', 'Move ID']
        for col, header in enumerate(headers, 1):
            cell = ws5.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font

        row = 2
        for err in self.vat_data['errori']:
            ws5.cell(row, 1, err['tipo'])
            ws5.cell(row, 2, err.get('data', ''))
            ws5.cell(row, 3, err.get('descrizione', ''))

            # Dettagli specifici per tipo errore
            dettagli = []
            if 'aliquota_calcolata' in err:
                dettagli.append(f"Aliquota: {err['aliquota_calcolata']}")
            if 'base_imponibile' in err:
                dettagli.append(f"Base: CHF {err['base_imponibile']:.2f}")
            if 'iva' in err:
                dettagli.append(f"IVA: CHF {err['iva']:.2f}")
            if 'fatture' in err:
                dettagli.append(f"Fatture: {', '.join(err['fatture'])}")

            ws5.cell(row, 4, ' | '.join(dettagli))
            ws5.cell(row, 5, err.get('move_id', ''))

            # Evidenzia in rosso
            ws5.cell(row, 1).fill = error_fill
            ws5.cell(row, 1).font = Font(color='FFFFFF')

            row += 1

        # 6. DETTAGLIO VENDITE
        ws6 = wb.create_sheet('Dettaglio Vendite')
        headers = ['Data', 'Mese', 'Descrizione', 'Partner', 'Base Imponibile', 'IVA', 'Aliquota', 'Move ID']
        for col, header in enumerate(headers, 1):
            cell = ws6.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font

        row = 2
        for det in sorted(self.vat_data['dettagli_vendite'], key=lambda x: x['data']):
            ws6.cell(row, 1, det['data'])
            ws6.cell(row, 2, det['mese'])
            ws6.cell(row, 3, det['descrizione'])
            ws6.cell(row, 4, det['partner'])
            ws6.cell(row, 5, det['base_imponibile'])
            ws6.cell(row, 6, det['iva'])
            ws6.cell(row, 7, det['aliquota'])
            ws6.cell(row, 8, det['move_id'])

            ws6.cell(row, 5).number_format = '#,##0.00'
            ws6.cell(row, 6).number_format = '#,##0.00'

            row += 1

        # 7. DETTAGLIO ACQUISTI
        ws7 = wb.create_sheet('Dettaglio Acquisti')
        for col, header in enumerate(headers, 1):
            cell = ws7.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font

        row = 2
        for det in sorted(self.vat_data['dettagli_acquisti'], key=lambda x: x['data']):
            ws7.cell(row, 1, det['data'])
            ws7.cell(row, 2, det['mese'])
            ws7.cell(row, 3, det['descrizione'])
            ws7.cell(row, 4, det['partner'])
            ws7.cell(row, 5, det['base_imponibile'])
            ws7.cell(row, 6, det['iva'])
            ws7.cell(row, 7, det['aliquota'])
            ws7.cell(row, 8, det['move_id'])

            ws7.cell(row, 5).number_format = '#,##0.00'
            ws7.cell(row, 6).number_format = '#,##0.00'

            row += 1

        # Autosize colonne
        for ws in wb:
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

        # Salva file
        filename = 'c:\\Users\\lapa\\Desktop\\Claude Code\\app-hub-platform\\RICONCILIAZIONE-IVA-2024.xlsx'
        wb.save(filename)
        print(f"\nReport salvato: {filename}")

        return filename

    def run(self):
        """Esegue analisi completa"""
        print("\n" + "="*80)
        print("RICONCILIAZIONE IVA 2024 - AVVIO")
        print("="*80)

        # 1. IVA vendite
        total_vendite = self.analyze_sales_vat()

        # 2. IVA acquisti
        total_acquisti = self.analyze_purchases_vat()

        # 3. Quadratura trimestrale
        quadratura = self.calculate_quarterly_summary()

        # 4. Identifica errori
        self.find_vat_errors()

        # 5. Genera report Excel
        filename = self.generate_excel_report(total_vendite, total_acquisti, quadratura)

        # Riepilogo finale
        print("\n" + "="*80)
        print("RIEPILOGO FINALE")
        print("="*80)
        print(f"\nTOTALE IVA VENDITE 2024:  CHF {total_vendite:>12,.2f}")
        print(f"TOTALE IVA ACQUISTI 2024: CHF {total_acquisti:>12,.2f}")
        print(f"{'-'*45}")
        saldo = total_vendite - total_acquisti
        print(f"SALDO IVA 2024:           CHF {saldo:>12,.2f}")

        if saldo > 0:
            print(f"\nSTATO: IVA A DEBITO (da versare)")
        else:
            print(f"\nSTATO: IVA A CREDITO")

        print(f"\nERRORI IDENTIFICATI: {len(self.vat_data['errori'])}")
        print(f"\nREPORT GENERATO: {filename}")

        # Salva anche JSON per debug
        json_file = 'c:\\Users\\lapa\\Desktop\\Claude Code\\app-hub-platform\\riconciliazione-iva-2024.json'
        with open(json_file, 'w', encoding='utf-8') as f:
            # Converti Decimal in float per JSON
            data_json = {
                'vendite_mensili': {
                    mese: {k: float(v) for k, v in aliquote.items()}
                    for mese, aliquote in self.vat_data['vendite_mensili'].items()
                },
                'acquisti_mensili': {
                    mese: {k: float(v) for k, v in aliquote.items()}
                    for mese, aliquote in self.vat_data['acquisti_mensili'].items()
                },
                'errori': self.vat_data['errori'],
                'totali': {
                    'vendite': float(total_vendite),
                    'acquisti': float(total_acquisti),
                    'saldo': float(saldo)
                }
            }
            json.dump(data_json, f, indent=2, ensure_ascii=False)

        print(f"Dati JSON salvati: {json_file}")

        return {
            'vendite': total_vendite,
            'acquisti': total_acquisti,
            'saldo': saldo,
            'errori': len(self.vat_data['errori']),
            'excel_file': filename,
            'json_file': json_file
        }

if __name__ == '__main__':
    try:
        reconciliation = OdooVATReconciliation()
        result = reconciliation.run()

        print("\n" + "="*80)
        print("ANALISI COMPLETATA CON SUCCESSO!")
        print("="*80)

    except Exception as e:
        print(f"\nERRORE: {e}")
        import traceback
        traceback.print_exc()
