"""
MASTER DASHBOARD GENERATOR - Riconciliazione 2024
Genera Excel completo con tutti i dati consolidati

Author: Process Automator
Date: 16 Novembre 2025
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_master_dashboard():
    """Crea Excel Master Dashboard per Riconciliazione 2024"""

    # Create workbook
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create sheets
    create_executive_summary(wb)
    create_asset_accounts(wb)
    create_technical_accounts(wb)
    create_action_plan(wb)
    create_financial_impact(wb)
    create_tracking(wb)

    # Save
    filename = f"MASTER-RICONCILIAZIONE-2024-{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    print(f"Dashboard creata: {filename}")

    return filename

def create_executive_summary(wb):
    """TAB 1: Executive Summary"""
    ws = wb.create_sheet("Executive Summary")

    # Title
    ws['A1'] = "MASTER DASHBOARD - RICONCILIAZIONE 2024"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:G1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Metadata
    ws['A2'] = f"Generato: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A3'] = "Database: lapadevadmin-lapa-v2-staging-2406-25408900"
    ws['A4'] = "Commercialista: Patrick Angstmann - PAGG Treuhand AG"

    # KPI Dashboard - Headers
    row = 6
    ws[f'A{row}'] = "KPI DASHBOARD"
    ws[f'A{row}'].font = Font(size=14, bold=True)

    row += 1
    headers = ['KPI', 'Baseline', 'Target', 'Status', 'Priority', 'Owner', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # KPI Data
    kpis = [
        ['Conti Tecnici Azzerati', '0/4 (0%)', '4/4 (100%)', 'RED 0%', 'P0', 'Accountant', 'CRITICO'],
        ['- Konto 1022', 'CHF 148,549', 'CHF 0.00', 'RED TODO', 'P0', 'Accountant', 'merenda69 blocker'],
        ['- Konto 1023', 'CHF -84,573', 'CHF 0.00', 'RED TODO', 'P0', 'Accountant', 'Fix XML-RPC'],
        ['- Konto 10901', 'CHF -375,616', 'CHF 0.00', 'RED TODO', 'P0', 'DB Optimizer', 'SQL pronte'],
        ['- Konto 1099', 'CHF -60,842', 'CHF 0.00', 'RED TODO', 'P1', 'Developer', 'Fix script'],
        ['Bank Accounts Aligned', '0/8 (0%)', '8/8 (100%)', 'RED 0%', 'P0', 'Business Analyst', '300% variance'],
        ['- Max Delta', 'CHF 188,840', '<CHF 1.00', 'RED TODO', 'P0', 'Accountant', 'Account 1026'],
        ['Cash Account', 'CHF 386,337', '~CHF 90,000', 'RED TODO', 'P1', 'Commercialista', 'BLOCKED'],
        ['- Rettifiche Approvate', '0/5', '5/5', 'RED BLOCKED', 'P1', 'Commercialista', 'Attesa approval'],
        ['Balance Sheet', 'NON BILANCIATO', 'BILANCIATO', 'RED TODO', 'P0', 'Data Analyst', 'CHF 3.17M diff'],
        ['Documentazione', '50+ files', 'Complete', 'GREEN DONE', '-', 'Team', 'Completato'],
        ['Scripts Pronti', '20+ scripts', 'Ready', 'GREEN DONE', '-', 'Developers', 'Completato'],
    ]

    row += 1
    for kpi_data in kpis:
        for col, value in enumerate(kpi_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            # Color coding for status
            if 'RED' in str(value):
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif 'GREEN' in str(value):
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif 'BLOCKED' in str(value):
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        row += 1

    # Financial Summary
    row += 2
    ws[f'A{row}'] = "FINANCIAL SUMMARY"
    ws[f'A{row}'].font = Font(size=14, bold=True)

    row += 1
    headers = ['Categoria', 'Delta CHF', '% del Totale', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    financial_data = [
        ['Conti Tecnici (1022/1023/10901/1099)', 'CHF 669,580', '60.3%', 'DA AZZERARE'],
        ['Cash Rettifiche', 'CHF 174,290', '15.7%', 'DA STORNARE'],
        ['Bank Reconciliation', 'CHF 345,458', '31.1%', 'DA ALLINEARE'],
        ['Cash Duplicati', 'CHF 784', '0.1%', 'DA ELIMINARE'],
        ['TOTALE CORREZIONI', 'CHF 1,190,112', '107.2%', 'IN PROGRESS'],
    ]

    row += 1
    for data in financial_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            if 'TOTALE' in str(data[0]):
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 25

def create_asset_accounts(wb):
    """TAB 2: Asset Accounts"""
    ws = wb.create_sheet("Asset Accounts")

    # Title
    ws['A1'] = "ASSET ACCOUNTS - DISCREPANZE"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:H1')

    # Cash 1001
    row = 3
    ws[f'A{row}'] = "KONTO 1001 - Cash"
    ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws.merge_cells(f'A{row}:H{row}')

    row += 1
    ws[f'A{row}'] = "Saldo Attuale:"
    ws[f'B{row}'] = "CHF 386,336.67"
    ws[f'C{row}'] = "Saldo Target:"
    ws[f'D{row}'] = "~CHF 90,000"
    ws[f'E{row}'] = "Delta:"
    ws[f'F{row}'] = "CHF 296,336.67"

    row += 2
    headers = ['Data', 'ID', 'Descrizione', 'Importo', '% Saldo', 'Tipo', 'Azione', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    rettifiche_data = [
        ['31.12.2023', '525905', 'Rettifica Cash da 21.396,03 a 109.280,46', 'CHF 87,884.43', '22.7%',
         'Rettifica sospetta', 'STORNARE', 'RED ATTESA APPROVAZIONE'],
        ['31.01.2024', '525812', 'Rettifica in aumento saldo 1000 - Cash', 'CHF 86,405.83', '22.4%',
         'Rettifica sospetta', 'STORNARE', 'RED ATTESA APPROVAZIONE'],
        ['01.10.2025', '523317/522654', 'Deposito - Rusu stefan → Nuraghets', 'CHF 400.00', '0.1%',
         'Duplicato', 'ELIMINARE', 'RED TODO'],
        ['13.02.2024', '234764/234762', 'Ordine 02030-003-0006 - DL SERVICES', 'CHF 174.25', '0.05%',
         'Duplicato', 'ELIMINARE', 'RED TODO'],
        ['13.01.2024', '503096/115978', "Ordine 02611-001-0001 - EMMA'S CAFE", 'CHF 209.47', '0.05%',
         'Duplicato', 'ELIMINARE', 'RED TODO'],
    ]

    row += 1
    for data in rettifiche_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            if 'RED' in str(value):
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        row += 1

    # Bank Accounts
    row += 2
    ws[f'A{row}'] = "BANK ACCOUNTS - RICONCILIAZIONE"
    ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws.merge_cells(f'A{row}:H{row}')

    row += 2
    headers = ['Conto', 'Nome', 'Saldo Odoo', 'Saldo Reale 31.12.24', 'Delta', 'Variance', 'Severity', 'Root Cause']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    bank_data = [
        ['1026', 'UBS CHF Unternehmen', 'CHF 371,453.70', 'CHF 182,613.26', 'CHF +188,840.44', '103.4%',
         'RED CRITICAL', 'Mapping error'],
        ['1024', 'UBS Privatkonto', 'CHF 121,554.65', 'CHF 23,783.88', 'CHF +97,770.77', '411.1%',
         'RED CRITICAL', 'Mapping error'],
        ['1025', 'CS Hauptkonto', 'CHF 108,267.67', 'CHF 11,120.67', 'CHF +97,147.00', '873.6%',
         'RED CRITICAL', 'Mapping error'],
        ['1021', 'UBS COVID', 'CHF -154,149.93', 'CHF -116,500.00', 'CHF -37,649.93', '32.3%',
         'ORANGE HIGH', 'Timing differences'],
        ['1027', 'CS Zweitkonto', 'CHF 13,032.22', 'CHF 13,777.05', 'CHF -744.83', '-5.4%',
         'YELLOW MEDIUM', 'Commissioni Q4'],
        ['1028', 'UBS EUR', 'CHF -1,340.43', 'EUR 128,860.70', 'FX ERROR', 'N/A',
         'RED CRITICAL', 'FX conversion'],
        ['1029', 'UBS USD', 'CHF -997.28', 'USD 92.63', 'FX ERROR', 'N/A',
         'RED CRITICAL', 'FX conversion'],
        ['1034', 'UNKNOWN', 'CHF 94.26', 'CHF 0.00', 'CHF +94.26', 'N/A',
         'ORANGE HIGH', 'Unmapped account'],
    ]

    row += 1
    for data in bank_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            if 'RED' in str(value):
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif 'ORANGE' in str(value):
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif 'YELLOW' in str(value):
                cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        row += 1

    # Adjust widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 25

def create_technical_accounts(wb):
    """TAB 3: Technical Accounts"""
    ws = wb.create_sheet("Technical Accounts")

    # Title
    ws['A1'] = "TECHNICAL ACCOUNTS - DA AZZERARE"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:G1')

    row = 3
    headers = ['Account', 'Descrizione', 'Saldo Attuale', 'Target', 'Delta', 'Priority', 'Status', 'Owner']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    tech_accounts_data = [
        ['1022', 'Outstanding Receipts', 'CHF 148,549.24', 'CHF 0.00', 'CHF 148,549.24', 'P0',
         'RED BLOCKED', 'Accountant + Commercialista'],
        ['1023', 'Outstanding Payments', 'CHF -84,573.31', 'CHF 0.00', 'CHF 84,573.31', 'P0',
         'RED TODO', 'Accountant'],
        ['10901', 'Liquiditätstransfer', 'CHF -375,615.65', 'CHF 0.00', 'CHF 375,615.65', 'P0',
         'RED IN PROGRESS', 'DB Optimizer'],
        ['1099', 'Transferkonto', 'CHF -60,842.41', 'CHF 0.00', 'CHF 60,842.41', 'P1',
         'RED TODO', 'Developer + Accountant'],
    ]

    row += 1
    for data in tech_accounts_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            if 'RED' in str(value):
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        row += 1

    # 10901 Categorization
    row += 2
    ws[f'A{row}'] = "KONTO 10901 - CATEGORIZZAZIONE DETTAGLIATA"
    ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    ws.merge_cells(f'A{row}:H{row}')

    row += 2
    headers = ['Categoria', 'Count', 'Balance CHF', 'Priority', 'Target Account', 'Action', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    categorization_data = [
        ['Currency Exchange FX', '40', 'CHF -599,376.20', 'HIGH', '2660', 'Execute SQL', 'YELLOW READY'],
        ['Credit Card Payment', '15', 'CHF +44,144.51', 'HIGH', '10803', 'Execute SQL', 'YELLOW READY'],
        ['Bank Transfer Internal', '29', 'CHF +3,000.00', 'HIGH', 'Manual', 'Review each', 'RED TODO'],
        ['Currency Diff', '39', 'CHF +372,214.97', 'MEDIUM', '2660 or verify', 'Classify', 'RED TODO'],
        ['Instant Payment', '69', 'CHF -470,000.00', 'CRITICAL', 'Check duplicates!', 'VERIFY FIRST', 'RED TODO'],
        ['Cash Deposit', '4', 'CHF -87,570.00', 'MEDIUM', '1000', 'Execute SQL', 'YELLOW READY'],
        ['Other', '157', 'CHF +361,971.07', 'LOW', 'Manual', 'Categorize', 'RED TODO'],
    ]

    row += 1
    for data in categorization_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            if 'RED' in str(value):
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif 'YELLOW' in str(value):
                cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            if 'CRITICAL' in str(data[3]):
                cell.font = Font(bold=True, color="C00000")
        row += 1

    # Adjust widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15

def create_action_plan(wb):
    """TAB 4: Action Plan"""
    ws = wb.create_sheet("Action Plan")

    # Title
    ws['A1'] = "ACTION PLAN - RICONCILIAZIONE 2024"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:I1')

    # Quick Wins
    row = 3
    ws[f'A{row}'] = "QUICK WINS (Week 1: 16-18 Nov)"
    ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    ws.merge_cells(f'A{row}:I{row}')

    row += 1
    headers = ['ID', 'Task', 'Impact', 'Effort (h)', 'Owner', 'Start Date', 'End Date', 'Status', 'Blockers']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    quick_wins_data = [
        ['QW-1', 'Fix Technical Scripts', 'Unlock automations', '4', 'Developer', '16 Nov', '16 Nov', 'RED TODO', '-'],
        ['QW-2', 'Execute SQL 10901 (FX+CC)', 'CHF -555,232 riclass', '4', 'DB Optimizer', '16 Nov', '16 Nov', 'RED TODO', '-'],
        ['QW-3', 'Chiusura Konto 1099', 'CHF 60,842 azzerato', '2', 'Accountant', '17 Nov', '17 Nov', 'RED TODO', 'QW-1'],
        ['QW-4', 'FX Conversion Banks', 'Clarify CHF -122K FX', '2', 'Business Analyst', '17 Nov', '17 Nov', 'RED TODO', '-'],
    ]

    row += 1
    for data in quick_wins_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
        row += 1

    # Medium Effort
    row += 1
    ws[f'A{row}'] = "MEDIUM EFFORT (Week 2: 19-23 Nov)"
    ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    ws.merge_cells(f'A{row}:I{row}')

    row += 1
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    medium_effort_data = [
        ['ME-1', 'Riconcilia 1022 (merenda69)', 'CHF 148,549 azzerato', '24', 'Accountant', '19 Nov', '21 Nov',
         'RED BLOCKED', 'Commercialista approval'],
        ['ME-2', 'Riconcilia 1023', 'CHF 84,573 azzerato', '16', 'Accountant', '22 Nov', '23 Nov',
         'RED TODO', 'QW-1 (fix script)'],
    ]

    row += 1
    for data in medium_effort_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            if 'BLOCKED' in str(value):
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        row += 1

    # Long Term
    row += 1
    ws[f'A{row}'] = "LONG TERM (Week 3-4: 26 Nov - 6 Dic)"
    ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws.merge_cells(f'A{row}:I{row}')

    row += 1
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    long_term_data = [
        ['LT-1', 'Bank Reconciliation (8 accounts)', 'CHF 345,458 aligned', '37', 'Business Analyst',
         '26 Nov', '28 Nov', 'RED TODO', 'QW-4'],
        ['LT-2', 'Complete 10901 Riclass', 'CHF 375,616 azzerato', '27', 'DB Optimizer',
         '29 Nov', '3 Dic', 'RED TODO', 'QW-2'],
        ['LT-3', 'Cash 1001 Rettifiche', 'CHF 175,074 corretto', '8', 'Accountant',
         '4 Dic', '4 Dic', 'RED BLOCKED', 'Commercialista approval'],
    ]

    row += 1
    for data in long_term_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            if 'BLOCKED' in str(value):
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        row += 1

    # Summary
    row += 2
    ws[f'A{row}'] = "TOTALE EFFORT: 132 ore | TARGET: 6 Dicembre 2025"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    ws.merge_cells(f'A{row}:I{row}')

    # Adjust widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 25

def create_financial_impact(wb):
    """TAB 5: Financial Impact"""
    ws = wb.create_sheet("Financial Impact")

    ws['A1'] = "FINANCIAL IMPACT - BEFORE/AFTER"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:F1')

    row = 3
    ws[f'A{row}'] = "BALANCE SHEET - PRIMA DELLE CORREZIONI"
    ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws.merge_cells(f'A{row}:F{row}')

    row += 2
    before_data = [
        ['ASSETS', 'CHF 1,793,244.60'],
        ['LIABILITIES', 'CHF -675,706.81'],
        ['EQUITY', 'CHF -702,779.98'],
        ['BALANCE CHECK', 'CHF 3,171,731.39 ❌ NON BILANCIATO'],
    ]

    for data in before_data:
        ws[f'A{row}'] = data[0]
        ws[f'B{row}'] = data[1]
        if 'NON BILANCIATO' in data[1]:
            ws[f'B{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            ws[f'B{row}'].font = Font(bold=True, color="C00000")
        row += 1

    row += 2
    ws[f'A{row}'] = "BALANCE SHEET - DOPO CORREZIONI (STIMATO)"
    ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    ws.merge_cells(f'A{row}:F{row}')

    row += 2
    after_data = [
        ['ASSETS', 'CHF 1,447,787 (dopo rettifiche Cash e bank alignment)'],
        ['LIABILITIES', 'CHF -675,707 (invariato)'],
        ['EQUITY', 'CHF -702,780 + P&L adjustments'],
        ['BALANCE CHECK', 'DA VERIFICARE DOPO CORREZIONI'],
    ]

    for data in after_data:
        ws[f'A{row}'] = data[0]
        ws[f'B{row}'] = data[1]
        row += 1

    # Corrections detail
    row += 2
    ws[f'A{row}'] = "DETTAGLIO CORREZIONI"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    ws.merge_cells(f'A{row}:F{row}')

    row += 2
    headers = ['Correzione', 'Account', 'Amount CHF', 'Type', 'Impact on BS']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    corrections_data = [
        ['Azzeramento 1022', '1022', 'CHF -148,549', 'Reconciliation', 'Assets decrease'],
        ['Azzeramento 1023', '1023', 'CHF +84,573', 'Reconciliation', 'Assets increase'],
        ['Azzeramento 10901', '10901', 'CHF +375,616', 'Reclassification', 'No impact (riclass)'],
        ['Azzeramento 1099', '1099', 'CHF +60,842', 'Transfer to Equity', 'Equity increase'],
        ['Storno Rettifiche Cash', '1001', 'CHF -174,290', 'Reversal', 'Assets decrease'],
        ['Storno Duplicati Cash', '1001', 'CHF -784', 'Removal', 'Assets decrease'],
        ['Bank Alignment', '1021-1034', 'CHF -345,458', 'Reconciliation', 'Assets decrease'],
    ]

    row += 1
    for data in corrections_data:
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 20

def create_tracking(wb):
    """TAB 6: Tracking"""
    ws = wb.create_sheet("Tracking")

    ws['A1'] = "TRACKING & MONITORING"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:H1')

    row = 3
    ws[f'A{row}'] = "OVERALL PROGRESS"
    ws[f'A{row}'].font = Font(size=12, bold=True)

    row += 2
    headers = ['Metric', 'Current', 'Target', '% Complete']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    progress_data = [
        ['Tasks Completed', '0', '9', '0%'],
        ['Hours Logged', '0', '132', '0%'],
        ['Amount Corrected', 'CHF 0', 'CHF 1,190,112', '0%'],
        ['Accounts Closed', '0/4 technical', '4/4', '0%'],
        ['Banks Aligned', '0/8', '8/8', '0%'],
        ['Approvals Received', '0/2', '2/2 (commercialista)', '0%'],
    ]

    row += 1
    for data in progress_data:
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    # Issues Log
    row += 2
    ws[f'A{row}'] = "ISSUES LOG"
    ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws.merge_cells(f'A{row}:H{row}')

    row += 2
    headers = ['ID', 'Date', 'Issue', 'Severity', 'Owner', 'Status', 'Resolution', 'Closed Date']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    issues_data = [
        ['ISS-001', '16 Nov', 'Commercialista approval needed - Cash rettifiche', 'CRITICAL',
         'Project Manager', 'OPEN', '-', '-'],
        ['ISS-002', '16 Nov', 'merenda69 CHF 182K - unknown origin', 'CRITICAL',
         'Commercialista', 'OPEN', '-', '-'],
        ['ISS-003', '16 Nov', 'Bank statements missing', 'HIGH',
         'Accountant', 'OPEN', '-', '-'],
    ]

    row += 1
    for data in issues_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            if 'CRITICAL' in str(value):
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        row += 1

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 12

if __name__ == "__main__":
    print("Generating Master Dashboard Excel...")
    filename = create_master_dashboard()
    print(f"DONE! File: {filename}")
    print("\nDashboard includes:")
    print("  1. Executive Summary (KPI Dashboard)")
    print("  2. Asset Accounts (Cash + Banks)")
    print("  3. Technical Accounts (1022/1023/10901/1099)")
    print("  4. Action Plan (Quick Wins, Medium, Long Term)")
    print("  5. Financial Impact (Before/After)")
    print("  6. Tracking (Progress, Issues)")
    print("\nReady to attach to email commercialista!")
