#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
RICONCILIAZIONE CONTI PATRIMONIALI (ASSET) 2024
Database Optimizer - Performance-focused reconciliation

Analizza TUTTI i conti patrimoniali (1000-1599) per trovare discrepanze
tra saldo teorico e saldo reale.
"""

import sys
import io
import xmlrpc.client
import os
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json

# Fix Windows encoding
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Odoo Connection
ODOO_URL = "https://lapadevadmin-lapa-v2-staging-2406-25408900.dev.odoo.com"
ODOO_DB = "lapadevadmin-lapa-v2-staging-2406-25408900"
ODOO_USER = "paul@lapa.ch"
ODOO_PASSWORD = "lapa201180"

# Account ranges to analyze (ASSET accounts)
ASSET_RANGES = {
    "1000-1099": "Liquidità (Cash, Banche, Transfer)",
    "1100-1199": "Crediti Clienti (Accounts Receivable)",
    "1200-1299": "Inventario/Magazzino (Inventory)",
    "1300-1399": "Crediti Diversi",
    "1500-1599": "Immobilizzazioni (Fixed Assets)"
}

class OdooConnection:
    """Optimized Odoo connection with batch operations"""

    def __init__(self):
        self.common = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/common')
        self.models = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/object')
        self.uid = None

    def connect(self):
        """Authenticate with Odoo"""
        print("🔌 Connessione a Odoo...")
        self.uid = self.common.authenticate(ODOO_DB, ODOO_USER, ODOO_PASSWORD, {})
        if not self.uid:
            raise Exception("Autenticazione fallita!")
        print(f"✅ Connesso come UID: {self.uid}")
        return self.uid

    def execute(self, model, method, domain, params):
        """Execute Odoo RPC call"""
        return self.models.execute_kw(
            ODOO_DB, self.uid, ODOO_PASSWORD,
            model, method, [domain], params
        )

    def search_read(self, model, domain, fields, **kwargs):
        """Optimized search_read with field selection"""
        params = {'fields': fields}
        params.update(kwargs)
        return self.execute(model, 'search_read', domain, params)

class AssetReconciliation:
    """Main reconciliation engine"""

    def __init__(self, odoo):
        self.odoo = odoo
        self.accounts_cache = {}
        self.move_lines_cache = {}
        self.discrepancies = []

    def get_accounts_in_range(self, code_start, code_end):
        """Get all accounts in code range (optimized query)"""
        cache_key = f"{code_start}-{code_end}"
        if cache_key in self.accounts_cache:
            return self.accounts_cache[cache_key]

        print(f"  📊 Fetching accounts {code_start}-{code_end}...")

        # Optimized domain with code range
        domain = [
            ('code', '>=', code_start),
            ('code', '<=', code_end),
            ('deprecated', '=', False)
        ]

        # Minimal field selection
        fields = ['id', 'code', 'name', 'account_type', 'reconcile']

        accounts = self.odoo.search_read('account.account', domain, fields,
                                        order='code asc')

        self.accounts_cache[cache_key] = accounts
        print(f"  ✅ Found {len(accounts)} accounts")
        return accounts

    def get_opening_balance(self, account_id, account_code):
        """Get opening balance as of 2024-01-01"""
        # Query ALL moves before 2024-01-01
        domain = [
            ('account_id', '=', account_id),
            ('date', '<', '2024-01-01'),
            ('parent_state', '=', 'posted')  # Only posted moves
        ]

        fields = ['debit', 'credit', 'balance']

        lines = self.odoo.search_read('account.move.line', domain, fields)

        # Calculate opening balance
        opening_debit = sum(line.get('debit', 0) for line in lines)
        opening_credit = sum(line.get('credit', 0) for line in lines)
        opening_balance = opening_debit - opening_credit

        return {
            'debit': opening_debit,
            'credit': opening_credit,
            'balance': opening_balance,
            'count': len(lines)
        }

    def get_moves_2024(self, account_id):
        """Get ALL 2024 moves for account (optimized batch query)"""
        cache_key = f"moves_{account_id}"
        if cache_key in self.move_lines_cache:
            return self.move_lines_cache[cache_key]

        domain = [
            ('account_id', '=', account_id),
            ('date', '>=', '2024-01-01'),
            ('date', '<=', '2024-12-31'),
            ('parent_state', '=', 'posted')
        ]

        # Essential fields only
        fields = [
            'id', 'date', 'name', 'ref', 'debit', 'credit',
            'balance', 'move_id', 'partner_id', 'amount_currency'
        ]

        moves = self.odoo.search_read('account.move.line', domain, fields,
                                     order='date asc, id asc')

        self.move_lines_cache[cache_key] = moves
        return moves

    def get_closing_balance(self, account_id):
        """Get actual closing balance as of 2024-12-31"""
        domain = [
            ('account_id', '=', account_id),
            ('date', '<=', '2024-12-31'),
            ('parent_state', '=', 'posted')
        ]

        fields = ['debit', 'credit', 'balance']

        lines = self.odoo.search_read('account.move.line', domain, fields)

        closing_debit = sum(line.get('debit', 0) for line in lines)
        closing_credit = sum(line.get('credit', 0) for line in lines)
        closing_balance = closing_debit - closing_credit

        return {
            'debit': closing_debit,
            'credit': closing_credit,
            'balance': closing_balance,
            'count': len(lines)
        }

    def analyze_account(self, account):
        """Complete analysis for single account"""
        account_id = account['id']
        account_code = account['code']
        account_name = account['name']

        print(f"\n🔍 Analyzing {account_code} - {account_name}...")

        # 1. Opening balance
        opening = self.get_opening_balance(account_id, account_code)
        print(f"  📌 Opening: {opening['balance']:,.2f} CHF ({opening['count']} moves)")

        # 2. 2024 moves
        moves_2024 = self.get_moves_2024(account_id)
        print(f"  📝 Moves 2024: {len(moves_2024)} lines")

        # 3. Calculate theoretical closing
        moves_debit = sum(m.get('debit', 0) for m in moves_2024)
        moves_credit = sum(m.get('credit', 0) for m in moves_2024)
        moves_balance = moves_debit - moves_credit

        theoretical_closing = opening['balance'] + moves_balance

        print(f"  📊 Moves D/C: {moves_debit:,.2f} / {moves_credit:,.2f}")
        print(f"  🧮 Theoretical closing: {theoretical_closing:,.2f} CHF")

        # 4. Actual closing balance
        actual = self.get_closing_balance(account_id)
        print(f"  ✔️  Actual closing: {actual['balance']:,.2f} CHF")

        # 5. Check discrepancy
        discrepancy = actual['balance'] - theoretical_closing

        result = {
            'account': account,
            'opening': opening,
            'moves_2024': {
                'count': len(moves_2024),
                'debit': moves_debit,
                'credit': moves_credit,
                'balance': moves_balance,
                'details': moves_2024
            },
            'theoretical_closing': theoretical_closing,
            'actual_closing': actual,
            'discrepancy': discrepancy,
            'has_discrepancy': abs(discrepancy) > 0.01  # Tolerance 1 cent
        }

        if result['has_discrepancy']:
            print(f"  ⚠️  DISCREPANCY: {discrepancy:,.2f} CHF")
            self.discrepancies.append(result)

            # Analyze discrepancy causes
            self.analyze_discrepancy(result)
        else:
            print(f"  ✅ OK - Balanced")

        return result

    def analyze_discrepancy(self, result):
        """Deep dive into discrepancy causes"""
        print(f"    🔬 Analyzing discrepancy causes...")

        moves = result['moves_2024']['details']

        # Check for duplicates (same date, same amount)
        duplicates = self.find_duplicates(moves)
        if duplicates:
            print(f"    🔴 Found {len(duplicates)} potential duplicates")

        # Check for large movements (outliers)
        large_moves = [m for m in moves if abs(m.get('debit', 0) - m.get('credit', 0)) > 10000]
        if large_moves:
            print(f"    💰 Found {len(large_moves)} large movements (>10k)")

        # Check for round numbers (suspicious)
        round_moves = [m for m in moves if self.is_round_number(m.get('debit', 0) + m.get('credit', 0))]
        if round_moves:
            print(f"    🔢 Found {len(round_moves)} round number movements")

        result['analysis'] = {
            'duplicates': duplicates,
            'large_moves': large_moves,
            'round_moves': round_moves
        }

    def find_duplicates(self, moves):
        """Find potential duplicate moves"""
        seen = {}
        duplicates = []

        for move in moves:
            key = (
                move.get('date'),
                round(move.get('debit', 0), 2),
                round(move.get('credit', 0), 2),
                move.get('partner_id', [False])[0] if move.get('partner_id') else False
            )

            if key in seen:
                duplicates.append({
                    'original': seen[key],
                    'duplicate': move
                })
            else:
                seen[key] = move

        return duplicates

    def is_round_number(self, amount):
        """Check if amount is suspiciously round"""
        if amount == 0:
            return False
        return amount % 100 == 0 or amount % 1000 == 0

    def reconcile_all_assets(self):
        """Main reconciliation workflow"""
        print("\n" + "="*70)
        print("🏦 RICONCILIAZIONE CONTI PATRIMONIALI (ASSET) 2024")
        print("="*70)

        all_results = []

        for range_code, description in ASSET_RANGES.items():
            print(f"\n{'='*70}")
            print(f"📂 {range_code}: {description}")
            print(f"{'='*70}")

            code_start, code_end = range_code.split('-')
            accounts = self.get_accounts_in_range(code_start, code_end)

            if not accounts:
                print("  ⚠️  No accounts found in this range")
                continue

            for account in accounts:
                result = self.analyze_account(account)
                all_results.append(result)

        return all_results

class ExcelReportGenerator:
    """Generate Excel report with reconciliation results"""

    def __init__(self, results):
        self.results = results
        self.wb = openpyxl.Workbook()

    def generate(self, filename):
        """Create complete Excel report"""
        print("\n📊 Generating Excel report...")

        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        # Create sheets
        self.create_summary_sheet()
        self.create_discrepancies_sheet()
        self.create_all_accounts_sheet()
        self.create_movements_sheet()

        # Save
        self.wb.save(filename)
        print(f"✅ Report saved: {filename}")

    def create_summary_sheet(self):
        """Summary dashboard"""
        ws = self.wb.create_sheet("Summary", 0)

        # Title
        ws['A1'] = "RICONCILIAZIONE ASSET 2024 - SUMMARY"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws.merge_cells('A1:F1')

        # Timestamp
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws.merge_cells('A2:F2')

        row = 4

        # Statistics
        ws[f'A{row}'] = "STATISTICS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        total_accounts = len(self.results)
        discrepant_accounts = len([r for r in self.results if r['has_discrepancy']])
        balanced_accounts = total_accounts - discrepant_accounts

        stats = [
            ("Total Accounts Analyzed", total_accounts),
            ("Accounts with Discrepancies", discrepant_accounts),
            ("Balanced Accounts", balanced_accounts),
            ("Discrepancy Rate", f"{(discrepant_accounts/total_accounts*100 if total_accounts > 0 else 0):.1f}%")
        ]

        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        row += 2

        # Discrepancy summary by range
        ws[f'A{row}'] = "DISCREPANCIES BY RANGE"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        headers = ['Range', 'Description', 'Accounts', 'Discrepancies', 'Total Discrepancy']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        row += 1

        for range_code, description in ASSET_RANGES.items():
            code_start, code_end = range_code.split('-')
            range_results = [r for r in self.results
                           if code_start <= r['account']['code'] <= code_end]

            if not range_results:
                continue

            discrepancies = [r for r in range_results if r['has_discrepancy']]
            total_disc = sum(r['discrepancy'] for r in discrepancies)

            ws.cell(row=row, column=1, value=range_code)
            ws.cell(row=row, column=2, value=description)
            ws.cell(row=row, column=3, value=len(range_results))
            ws.cell(row=row, column=4, value=len(discrepancies))
            ws.cell(row=row, column=5, value=total_disc).number_format = '#,##0.00'

            if len(discrepancies) > 0:
                ws.cell(row=row, column=4).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18

    def create_discrepancies_sheet(self):
        """Detailed discrepancies"""
        ws = self.wb.create_sheet("Discrepancies")

        # Headers
        headers = [
            'Account Code', 'Account Name', 'Opening Balance',
            'Moves Debit', 'Moves Credit', 'Moves Balance',
            'Theoretical Closing', 'Actual Closing', 'DISCREPANCY',
            'Duplicates', 'Large Moves', 'Status'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        row = 2
        discrepancies = [r for r in self.results if r['has_discrepancy']]

        for result in discrepancies:
            account = result['account']
            analysis = result.get('analysis', {})

            ws.cell(row=row, column=1, value=account['code'])
            ws.cell(row=row, column=2, value=account['name'])
            ws.cell(row=row, column=3, value=result['opening']['balance']).number_format = '#,##0.00'
            ws.cell(row=row, column=4, value=result['moves_2024']['debit']).number_format = '#,##0.00'
            ws.cell(row=row, column=5, value=result['moves_2024']['credit']).number_format = '#,##0.00'
            ws.cell(row=row, column=6, value=result['moves_2024']['balance']).number_format = '#,##0.00'
            ws.cell(row=row, column=7, value=result['theoretical_closing']).number_format = '#,##0.00'
            ws.cell(row=row, column=8, value=result['actual_closing']['balance']).number_format = '#,##0.00'

            disc_cell = ws.cell(row=row, column=9, value=result['discrepancy'])
            disc_cell.number_format = '#,##0.00'
            disc_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            disc_cell.font = Font(bold=True)

            ws.cell(row=row, column=10, value=len(analysis.get('duplicates', [])))
            ws.cell(row=row, column=11, value=len(analysis.get('large_moves', [])))
            ws.cell(row=row, column=12, value="NEEDS REVIEW")

            row += 1

        # Adjust widths
        for col in range(1, 13):
            ws.column_dimensions[get_column_letter(col)].width = 15
        ws.column_dimensions['B'].width = 35

    def create_all_accounts_sheet(self):
        """All accounts overview"""
        ws = self.wb.create_sheet("All Accounts")

        headers = [
            'Account Code', 'Account Name', 'Type', 'Opening Balance',
            'Moves Count', 'Moves Debit', 'Moves Credit',
            'Theoretical Closing', 'Actual Closing', 'Difference', 'Status'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        row = 2
        for result in self.results:
            account = result['account']

            ws.cell(row=row, column=1, value=account['code'])
            ws.cell(row=row, column=2, value=account['name'])
            ws.cell(row=row, column=3, value=account.get('account_type', 'N/A'))
            ws.cell(row=row, column=4, value=result['opening']['balance']).number_format = '#,##0.00'
            ws.cell(row=row, column=5, value=result['moves_2024']['count'])
            ws.cell(row=row, column=6, value=result['moves_2024']['debit']).number_format = '#,##0.00'
            ws.cell(row=row, column=7, value=result['moves_2024']['credit']).number_format = '#,##0.00'
            ws.cell(row=row, column=8, value=result['theoretical_closing']).number_format = '#,##0.00'
            ws.cell(row=row, column=9, value=result['actual_closing']['balance']).number_format = '#,##0.00'
            ws.cell(row=row, column=10, value=result['discrepancy']).number_format = '#,##0.00'

            status_cell = ws.cell(row=row, column=11, value="OK" if not result['has_discrepancy'] else "DISCREPANCY")
            if result['has_discrepancy']:
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                status_cell.font = Font(bold=True)
            else:
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

            row += 1

        for col in range(1, 12):
            ws.column_dimensions[get_column_letter(col)].width = 15
        ws.column_dimensions['B'].width = 35

    def create_movements_sheet(self):
        """Detailed movements for discrepant accounts"""
        ws = self.wb.create_sheet("Movements Detail")

        headers = [
            'Account Code', 'Account Name', 'Date', 'Move ID',
            'Name', 'Reference', 'Partner', 'Debit', 'Credit', 'Balance'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        row = 2
        discrepancies = [r for r in self.results if r['has_discrepancy']]

        for result in discrepancies:
            account = result['account']
            moves = result['moves_2024']['details']

            for move in moves:
                ws.cell(row=row, column=1, value=account['code'])
                ws.cell(row=row, column=2, value=account['name'])
                ws.cell(row=row, column=3, value=move.get('date', ''))
                ws.cell(row=row, column=4, value=move.get('move_id', [False, ''])[1] if move.get('move_id') else '')
                ws.cell(row=row, column=5, value=move.get('name', ''))
                ws.cell(row=row, column=6, value=move.get('ref', ''))
                ws.cell(row=row, column=7, value=move.get('partner_id', [False, ''])[1] if move.get('partner_id') else '')
                ws.cell(row=row, column=8, value=move.get('debit', 0)).number_format = '#,##0.00'
                ws.cell(row=row, column=9, value=move.get('credit', 0)).number_format = '#,##0.00'
                ws.cell(row=row, column=10, value=move.get('balance', 0)).number_format = '#,##0.00'

                row += 1

            # Add separator
            row += 1

        for col in range(1, 11):
            ws.column_dimensions[get_column_letter(col)].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['E'].width = 40

def main():
    """Main execution"""
    try:
        # Connect to Odoo
        odoo = OdooConnection()
        odoo.connect()

        # Run reconciliation
        reconciliation = AssetReconciliation(odoo)
        results = reconciliation.reconcile_all_assets()

        # Generate report
        print("\n" + "="*70)
        print("📊 GENERATING EXCEL REPORT")
        print("="*70)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'RICONCILIAZIONE-ASSET-2024-{timestamp}.xlsx'

        report = ExcelReportGenerator(results)
        report.generate(filename)

        # Summary
        print("\n" + "="*70)
        print("✅ RECONCILIATION COMPLETE")
        print("="*70)
        print(f"Total accounts analyzed: {len(results)}")
        print(f"Accounts with discrepancies: {len([r for r in results if r['has_discrepancy']])}")
        print(f"Report file: {filename}")

        # Save JSON for further analysis
        json_file = filename.replace('.xlsx', '.json')
        with open(json_file, 'w', encoding='utf-8') as f:
            # Simplify results for JSON (remove move details)
            simple_results = []
            for r in results:
                simple = {
                    'account_code': r['account']['code'],
                    'account_name': r['account']['name'],
                    'opening_balance': r['opening']['balance'],
                    'moves_count': r['moves_2024']['count'],
                    'moves_debit': r['moves_2024']['debit'],
                    'moves_credit': r['moves_2024']['credit'],
                    'theoretical_closing': r['theoretical_closing'],
                    'actual_closing': r['actual_closing']['balance'],
                    'discrepancy': r['discrepancy'],
                    'has_discrepancy': r['has_discrepancy']
                }
                simple_results.append(simple)

            json.dump(simple_results, f, indent=2, ensure_ascii=False)

        print(f"JSON export: {json_file}")

    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()
