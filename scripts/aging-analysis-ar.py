#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AGING ANALYSIS - ACCOUNTS RECEIVABLE (1100)
Database Optimizer - Performance-focused AR aging

Analizza i crediti clienti (1.1M CHF) per scadenze:
- 0-30 giorni
- 31-60 giorni
- 61-90 giorni
- 91-180 giorni
- 180+ giorni (critical)
"""

import sys
import io
import xmlrpc.client
from datetime import datetime, timedelta
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import json

# Fix Windows encoding
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Odoo Connection
ODOO_URL = "https://lapadevadmin-lapa-v2-staging-2406-25408900.dev.odoo.com"
ODOO_DB = "lapadevadmin-lapa-v2-staging-2406-25408900"
ODOO_USER = "paul@lapa.ch"
ODOO_PASSWORD = "lapa201180"

# Aging buckets (days)
AGING_BUCKETS = [
    (0, 30, "Current (0-30 days)"),
    (31, 60, "30-60 days"),
    (61, 90, "60-90 days"),
    (91, 180, "90-180 days"),
    (181, 99999, "180+ days (CRITICAL)")
]

class OdooConnection:
    def __init__(self):
        self.common = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/common')
        self.models = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/object')
        self.uid = None

    def connect(self):
        print("Connessione a Odoo...")
        self.uid = self.common.authenticate(ODOO_DB, ODOO_USER, ODOO_PASSWORD, {})
        if not self.uid:
            raise Exception("Autenticazione fallita!")
        print(f"Connesso come UID: {self.uid}")
        return self.uid

    def execute(self, model, method, domain, params):
        return self.models.execute_kw(
            ODOO_DB, self.uid, ODOO_PASSWORD,
            model, method, [domain], params
        )

class ARAgingAnalyzer:
    """Accounts Receivable Aging Analysis"""

    def __init__(self, odoo):
        self.odoo = odoo
        self.today = datetime.now().date()
        self.aging_data = defaultdict(lambda: defaultdict(float))
        self.partner_aging = defaultdict(lambda: defaultdict(float))

    def get_ar_account(self):
        """Get AR account (1100)"""
        print("\nFetching AR account...")
        accounts = self.odoo.execute('account.account', 'search_read',
            [('code', '=', '1100')],
            {'fields': ['id', 'code', 'name'], 'limit': 1}
        )

        if not accounts:
            raise Exception("AR account 1100 not found!")

        account = accounts[0]
        print(f"AR Account: {account['code']} - {account['name']}")
        return account

    def get_open_receivables(self, account_id):
        """Get all open (unreconciled) receivables"""
        print("\nFetching open receivables...")

        # Domain: AR account, not fully reconciled, posted
        domain = [
            ('account_id', '=', account_id),
            ('reconciled', '=', False),
            ('parent_state', '=', 'posted')
        ]

        # Essential fields for aging
        fields = [
            'id', 'date', 'date_maturity', 'name', 'ref',
            'partner_id', 'debit', 'credit', 'amount_residual',
            'amount_residual_currency', 'currency_id', 'move_id'
        ]

        lines = self.odoo.execute('account.move.line', 'search_read',
            domain,
            {'fields': fields, 'order': 'date asc'}
        )

        print(f"Found {len(lines)} open receivable lines")
        return lines

    def calculate_aging_bucket(self, line):
        """Calculate which aging bucket a line falls into"""
        # Use date_maturity if available, otherwise date
        due_date_str = line.get('date_maturity') or line.get('date')
        if not due_date_str:
            return "Unknown", 99999

        due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
        days_overdue = (self.today - due_date).days

        # Find appropriate bucket
        for min_days, max_days, label in AGING_BUCKETS:
            if min_days <= days_overdue <= max_days:
                return label, days_overdue

        return "Unknown", days_overdue

    def analyze_aging(self, lines):
        """Analyze aging for all receivables"""
        print("\nAnalyzing aging buckets...")

        results = []

        for line in lines:
            bucket, days_overdue = self.calculate_aging_bucket(line)
            amount = line.get('amount_residual', 0)

            # Skip if fully paid
            if abs(amount) < 0.01:
                continue

            partner_id = line.get('partner_id', [False, 'Unknown'])[0]
            partner_name = line.get('partner_id', [False, 'Unknown'])[1]

            result = {
                'id': line['id'],
                'date': line.get('date', ''),
                'date_maturity': line.get('date_maturity', ''),
                'move_id': line.get('move_id', [False, ''])[1],
                'name': line.get('name', ''),
                'ref': line.get('ref', ''),
                'partner_id': partner_id,
                'partner_name': partner_name,
                'amount': amount,
                'currency': line.get('currency_id', [False, 'CHF'])[1],
                'days_overdue': days_overdue,
                'aging_bucket': bucket
            }

            results.append(result)

            # Aggregate by bucket
            self.aging_data[bucket]['amount'] += amount
            self.aging_data[bucket]['count'] += 1

            # Aggregate by partner
            self.partner_aging[partner_name][bucket] += amount
            self.partner_aging[partner_name]['total'] += amount

        print(f"\nAging summary:")
        for bucket, data in sorted(self.aging_data.items()):
            print(f"  {bucket}: {data['amount']:,.2f} CHF ({data['count']} invoices)")

        return results

    def get_top_debtors(self, limit=20):
        """Get top N debtors by total amount"""
        sorted_partners = sorted(
            self.partner_aging.items(),
            key=lambda x: x[1]['total'],
            reverse=True
        )
        return sorted_partners[:limit]

class ExcelReportGenerator:
    def __init__(self, aging_results, aging_summary, top_debtors):
        self.results = aging_results
        self.summary = aging_summary
        self.top_debtors = top_debtors
        self.wb = openpyxl.Workbook()

    def generate(self, filename):
        print("\nGenerating Excel report...")

        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        self.create_summary_sheet()
        self.create_aging_detail_sheet()
        self.create_top_debtors_sheet()
        self.create_critical_sheet()

        self.wb.save(filename)
        print(f"Report saved: {filename}")

    def create_summary_sheet(self):
        ws = self.wb.create_sheet("Summary", 0)

        ws['A1'] = "AR AGING ANALYSIS - SUMMARY"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws.merge_cells('A1:D1')

        ws['A2'] = f"As of: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws.merge_cells('A2:D2')

        row = 4
        ws[f'A{row}'] = "AGING BUCKETS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        headers = ['Aging Bucket', 'Amount (CHF)', 'Count', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        row += 1

        total_amount = sum(data['amount'] for data in self.summary.values())

        for bucket in ["Current (0-30 days)", "30-60 days", "60-90 days", "90-180 days", "180+ days (CRITICAL)"]:
            data = self.summary.get(bucket, {'amount': 0, 'count': 0})
            percentage = (data['amount'] / total_amount * 100) if total_amount > 0 else 0

            ws.cell(row=row, column=1, value=bucket)
            ws.cell(row=row, column=2, value=data['amount']).number_format = '#,##0.00'
            ws.cell(row=row, column=3, value=data['count'])
            ws.cell(row=row, column=4, value=f"{percentage:.1f}%")

            # Highlight critical
            if bucket == "180+ days (CRITICAL)" and data['amount'] > 0:
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                    )

            row += 1

        # Total
        row += 1
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=2, value=total_amount).number_format = '#,##0.00'
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=3, value=sum(d['count'] for d in self.summary.values())).font = Font(bold=True)

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15

    def create_aging_detail_sheet(self):
        ws = self.wb.create_sheet("Aging Detail")

        headers = [
            'Partner', 'Invoice', 'Date', 'Due Date', 'Days Overdue',
            'Amount (CHF)', 'Aging Bucket', 'Reference', 'Name'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        row = 2
        for result in sorted(self.results, key=lambda x: x['days_overdue'], reverse=True):
            ws.cell(row=row, column=1, value=result['partner_name'])
            ws.cell(row=row, column=2, value=result['move_id'])
            ws.cell(row=row, column=3, value=result['date'])
            ws.cell(row=row, column=4, value=result['date_maturity'] or result['date'])
            ws.cell(row=row, column=5, value=result['days_overdue'])
            ws.cell(row=row, column=6, value=result['amount']).number_format = '#,##0.00'
            ws.cell(row=row, column=7, value=result['aging_bucket'])
            ws.cell(row=row, column=8, value=result['ref'] or '')
            ws.cell(row=row, column=9, value=result['name'] or '')

            # Highlight critical (180+ days)
            if result['days_overdue'] > 180:
                for col in range(1, 10):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                    )

            row += 1

        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 18
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['I'].width = 40

    def create_top_debtors_sheet(self):
        ws = self.wb.create_sheet("Top Debtors")

        ws['A1'] = "TOP 20 DEBTORS"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:G1')

        headers = [
            'Partner', 'Total (CHF)', '0-30 days', '31-60 days',
            '61-90 days', '91-180 days', '180+ days'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        row = 3
        for partner_name, aging in self.top_debtors:
            ws.cell(row=row, column=1, value=partner_name)
            ws.cell(row=row, column=2, value=aging['total']).number_format = '#,##0.00'

            buckets = [
                "Current (0-30 days)", "30-60 days", "60-90 days",
                "90-180 days", "180+ days (CRITICAL)"
            ]

            for col_idx, bucket in enumerate(buckets, 3):
                amount = aging.get(bucket, 0)
                cell = ws.cell(row=row, column=col_idx, value=amount)
                cell.number_format = '#,##0.00'

                # Highlight if significant amount in critical bucket
                if bucket == "180+ days (CRITICAL)" and amount > 1000:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            row += 1

        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 18
        ws.column_dimensions['A'].width = 35

    def create_critical_sheet(self):
        """Sheet with only critical (180+ days) receivables"""
        ws = self.wb.create_sheet("CRITICAL (180+ days)")

        ws['A1'] = "CRITICAL RECEIVABLES (180+ DAYS OVERDUE)"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        ws.merge_cells('A1:H1')

        headers = [
            'Partner', 'Invoice', 'Date', 'Due Date', 'Days Overdue',
            'Amount (CHF)', 'Reference', 'Name'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        critical = [r for r in self.results if r['days_overdue'] > 180]

        row = 3
        for result in sorted(critical, key=lambda x: x['amount'], reverse=True):
            ws.cell(row=row, column=1, value=result['partner_name'])
            ws.cell(row=row, column=2, value=result['move_id'])
            ws.cell(row=row, column=3, value=result['date'])
            ws.cell(row=row, column=4, value=result['date_maturity'] or result['date'])
            ws.cell(row=row, column=5, value=result['days_overdue'])
            ws.cell(row=row, column=6, value=result['amount']).number_format = '#,##0.00'
            ws.cell(row=row, column=7, value=result['ref'] or '')
            ws.cell(row=row, column=8, value=result['name'] or '')
            row += 1

        if not critical:
            ws['A3'] = "No critical receivables found!"
            ws['A3'].font = Font(color="00B050", bold=True)

        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 18
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['H'].width = 40

def main():
    try:
        odoo = OdooConnection()
        odoo.connect()

        analyzer = ARAgingAnalyzer(odoo)

        # Get AR account
        ar_account = analyzer.get_ar_account()

        # Get open receivables
        open_lines = analyzer.get_open_receivables(ar_account['id'])

        # Analyze aging
        aging_results = analyzer.analyze_aging(open_lines)

        # Get top debtors
        top_debtors = analyzer.get_top_debtors(20)

        print("\n" + "="*70)
        print("TOP 10 DEBTORS")
        print("="*70)
        for idx, (partner, aging) in enumerate(top_debtors[:10], 1):
            print(f"{idx:2d}. {partner:40s} {aging['total']:>15,.2f} CHF")

        # Generate report
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'AR-AGING-ANALYSIS-{timestamp}.xlsx'

        report = ExcelReportGenerator(
            aging_results,
            analyzer.aging_data,
            top_debtors
        )
        report.generate(filename)

        # JSON export
        json_file = filename.replace('.xlsx', '.json')
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump({
                'summary': dict(analyzer.aging_data),
                'results': aging_results,
                'top_debtors': [(p, dict(a)) for p, a in top_debtors]
            }, f, indent=2, ensure_ascii=False, default=str)

        print(f"\nJSON export: {json_file}")

        print("\n" + "="*70)
        print("AR AGING ANALYSIS COMPLETE")
        print("="*70)

    except Exception as e:
        print(f"\nERROR: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()
