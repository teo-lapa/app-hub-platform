#!/usr/bin/env python3
"""
Genera Excel con transazioni settembre 2024 per facile revisione
"""

import json
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed")
    print("Install with: pip install openpyxl")
    exit(1)

def create_excel_report():
    """Crea report Excel settembre 2024"""

    base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    report_file = os.path.join(base_path, 'REPORT-SETTEMBRE-2024.json')

    # Carica report
    with open(report_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Crea workbook
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # ========================================================================
    # SHEET 1: SUMMARY
    # ========================================================================
    ws_summary = wb.create_sheet('SUMMARY')

    # Header
    ws_summary['A1'] = 'SETTEMBRE 2024 - VERIFICA ESTRATTI BANCARI'
    ws_summary['A1'].font = Font(size=16, bold=True)
    ws_summary.merge_cells('A1:F1')

    ws_summary['A2'] = f"Generato: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    # Table header
    headers = ['Konto', 'Banca', 'Valuta', 'Saldo Inizio', 'Saldo Fine', 'Variazione', 'Transazioni']
    for col, header in enumerate(headers, start=1):
        cell = ws_summary.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    row = 5

    # 1024 CHF
    bank_1024 = data['accounts']['1024']['bank_statement']
    ws_summary[f'A{row}'] = '1024'
    ws_summary[f'B{row}'] = 'UBS'
    ws_summary[f'C{row}'] = 'CHF'
    ws_summary[f'D{row}'] = float(bank_1024['saldo_01_09'])
    ws_summary[f'E{row}'] = float(bank_1024['saldo_30_09'])
    ws_summary[f'F{row}'] = float(bank_1024['variazione'])
    ws_summary[f'G{row}'] = bank_1024['transactions_q3']

    # Format numbers
    for col in ['D', 'E', 'F']:
        ws_summary[f'{col}{row}'].number_format = '#,##0.00'

    row += 1

    # 1025 EUR
    bank_1025 = data['accounts']['1025']['bank_statement']
    ws_summary[f'A{row}'] = '1025'
    ws_summary[f'B{row}'] = 'UBS'
    ws_summary[f'C{row}'] = 'EUR'
    ws_summary[f'D{row}'] = float(bank_1025['saldo_01_09'])
    ws_summary[f'E{row}'] = float(bank_1025['saldo_30_09'])
    ws_summary[f'F{row}'] = float(bank_1025['variazione_saldi'])
    ws_summary[f'G{row}'] = bank_1025['num_transactions']

    # Format numbers
    for col in ['D', 'E', 'F']:
        ws_summary[f'{col}{row}'].number_format = '#,##0.00'

    # Highlight negative variation
    if float(bank_1025['variazione_saldi']) < 0:
        ws_summary[f'F{row}'].font = Font(color='FF0000')

    row += 1

    # 1026 CHF
    ws_summary[f'A{row}'] = '1026'
    ws_summary[f'B{row}'] = 'Credit Suisse'
    ws_summary[f'C{row}'] = 'CHF'
    ws_summary[f'D{row}'] = 'N/A'
    ws_summary[f'E{row}'] = 'N/A'
    ws_summary[f'F{row}'] = 'N/A'
    ws_summary[f'G{row}'] = 0

    # Column widths
    ws_summary.column_dimensions['A'].width = 10
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['C'].width = 10
    ws_summary.column_dimensions['D'].width = 15
    ws_summary.column_dimensions['E'].width = 15
    ws_summary.column_dimensions['F'].width = 15
    ws_summary.column_dimensions['G'].width = 15

    # ========================================================================
    # SHEET 2: EUR TRANSACTIONS
    # ========================================================================
    ws_eur = wb.create_sheet('EUR-1025-Transactions')

    # Header
    ws_eur['A1'] = 'KONTO 1025 - EUR UBS - SETTEMBRE 2024'
    ws_eur['A1'].font = Font(size=14, bold=True)
    ws_eur.merge_cells('A1:H1')

    # Table header
    headers_eur = ['Data', 'Partner', 'Descrizione', 'Dare', 'Avere', 'Saldo', 'Valuta', 'Note']
    for col, header in enumerate(headers_eur, start=1):
        cell = ws_eur.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    row = 4
    transactions = bank_1025['transactions']

    # Sort by date
    transactions_sorted = sorted(transactions, key=lambda x: x['date'])

    total_debit = 0
    total_credit = 0

    for tx in transactions_sorted:
        ws_eur[f'A{row}'] = tx['date']
        ws_eur[f'B{row}'] = tx.get('partner_name', '')[:30]
        ws_eur[f'C{row}'] = tx['description'][:100]

        amount = float(tx['amount'])
        if amount < 0:
            ws_eur[f'D{row}'] = abs(amount)
            total_debit += abs(amount)
        else:
            ws_eur[f'E{row}'] = amount
            total_credit += amount

        ws_eur[f'F{row}'] = float(tx.get('balance', 0))
        ws_eur[f'G{row}'] = tx.get('currency', 'EUR')

        # Note per operazioni FX
        if 'FX Forward' in tx['description']:
            ws_eur[f'H{row}'] = 'OPERAZIONE CAMBIO'
            ws_eur[f'H{row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # Format
        ws_eur[f'D{row}'].number_format = '#,##0.00'
        ws_eur[f'E{row}'].number_format = '#,##0.00'
        ws_eur[f'F{row}'].number_format = '#,##0.00'

        # Highlight negative balance
        if float(tx.get('balance', 0)) < 0:
            ws_eur[f'F{row}'].font = Font(color='FF0000')

        row += 1

    # TOTALS
    row += 1
    ws_eur[f'B{row}'] = 'TOTALE'
    ws_eur[f'B{row}'].font = Font(bold=True)
    ws_eur[f'D{row}'] = total_debit
    ws_eur[f'E{row}'] = total_credit
    ws_eur[f'D{row}'].number_format = '#,##0.00'
    ws_eur[f'E{row}'].number_format = '#,##0.00'
    ws_eur[f'D{row}'].font = Font(bold=True)
    ws_eur[f'E{row}'].font = Font(bold=True)

    row += 1
    ws_eur[f'B{row}'] = 'NETTO'
    ws_eur[f'B{row}'].font = Font(bold=True)
    ws_eur[f'F{row}'] = total_credit - total_debit
    ws_eur[f'F{row}'].number_format = '#,##0.00'
    ws_eur[f'F{row}'].font = Font(bold=True, color='FF0000' if (total_credit - total_debit) < 0 else '00FF00')

    # Column widths
    ws_eur.column_dimensions['A'].width = 12
    ws_eur.column_dimensions['B'].width = 35
    ws_eur.column_dimensions['C'].width = 50
    ws_eur.column_dimensions['D'].width = 15
    ws_eur.column_dimensions['E'].width = 15
    ws_eur.column_dimensions['F'].width = 15
    ws_eur.column_dimensions['G'].width = 10
    ws_eur.column_dimensions['H'].width = 25

    # ========================================================================
    # SHEET 3: ODOO CHECKLIST
    # ========================================================================
    ws_odoo = wb.create_sheet('ODOO-Checklist')

    ws_odoo['A1'] = 'ODOO VERIFICATION CHECKLIST'
    ws_odoo['A1'].font = Font(size=14, bold=True)
    ws_odoo.merge_cells('A1:D1')

    row = 3
    ws_odoo[f'A{row}'] = 'Konto'
    ws_odoo[f'B{row}'] = 'Descrizione'
    ws_odoo[f'C{row}'] = 'Saldo Banca 30/09'
    ws_odoo[f'D{row}'] = 'Saldo Odoo 30/09'
    ws_odoo[f'E{row}'] = 'Differenza'
    ws_odoo[f'F{row}'] = 'Status'

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_odoo[f'{col}{row}'].font = Font(bold=True)
        ws_odoo[f'{col}{row}'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        ws_odoo[f'{col}{row}'].font = Font(color='FFFFFF', bold=True)

    row += 1

    # 1024
    ws_odoo[f'A{row}'] = '1024'
    ws_odoo[f'B{row}'] = 'CHF-UBS PRINCIPALE'
    ws_odoo[f'C{row}'] = float(bank_1024['saldo_30_09'])
    ws_odoo[f'D{row}'] = ''  # To be filled
    ws_odoo[f'E{row}'].value = '=D{0}-C{0}'.format(row)
    ws_odoo[f'F{row}'] = 'DA VERIFICARE'

    row += 1

    # 1025
    ws_odoo[f'A{row}'] = '1025'
    ws_odoo[f'B{row}'] = 'EUR-UBS'
    ws_odoo[f'C{row}'] = float(bank_1025['saldo_30_09'])
    ws_odoo[f'D{row}'] = ''  # To be filled
    ws_odoo[f'E{row}'].value = '=D{0}-C{0}'.format(row)
    ws_odoo[f'F{row}'] = 'DA VERIFICARE'

    row += 1

    # 1026
    ws_odoo[f'A{row}'] = '1026'
    ws_odoo[f'B{row}'] = 'CHF-CRS PRINCIPALE'
    ws_odoo[f'C{row}'] = 'N/A'
    ws_odoo[f'D{row}'] = ''  # To be filled
    ws_odoo[f'E{row}'] = 'N/A'
    ws_odoo[f'F{row}'] = 'INCOMPLETO'

    # Format
    for r in range(4, row + 1):
        if ws_odoo[f'C{r}'].value != 'N/A':
            ws_odoo[f'C{r}'].number_format = '#,##0.00'
        ws_odoo[f'D{r}'].number_format = '#,##0.00'
        ws_odoo[f'E{r}'].number_format = '#,##0.00'

    # Column widths
    ws_odoo.column_dimensions['A'].width = 10
    ws_odoo.column_dimensions['B'].width = 30
    ws_odoo.column_dimensions['C'].width = 20
    ws_odoo.column_dimensions['D'].width = 20
    ws_odoo.column_dimensions['E'].width = 15
    ws_odoo.column_dimensions['F'].width = 20

    # Instructions
    row += 3
    ws_odoo[f'A{row}'] = 'ISTRUZIONI:'
    ws_odoo[f'A{row}'].font = Font(bold=True)
    row += 1
    ws_odoo[f'A{row}'] = '1. Eseguire query Odoo per ottenere saldi al 30/09/2024'
    row += 1
    ws_odoo[f'A{row}'] = '2. Inserire i valori nella colonna "Saldo Odoo 30/09"'
    row += 1
    ws_odoo[f'A{row}'] = '3. La differenza sarà calcolata automaticamente'
    row += 1
    ws_odoo[f'A{row}'] = '4. Se differenza < 1 EUR: OK, altrimenti investigare'

    # ========================================================================
    # SAVE
    # ========================================================================
    output_file = os.path.join(base_path, 'SETTEMBRE-2024-VERIFICA.xlsx')
    wb.save(output_file)

    print(f"\n[OK] Excel creato: {output_file}")
    print(f"\nSheet create:")
    print(f"  1. SUMMARY - Riepilogo konti")
    print(f"  2. EUR-1025-Transactions - 38 transazioni settembre 2024")
    print(f"  3. ODOO-Checklist - Checklist verifica Odoo")

    return output_file

if __name__ == '__main__':
    create_excel_report()
