#!/usr/bin/env python3
"""
RICONCILIAZIONE BANCARIA COMPLETA 2024 - MESE PER MESE
=======================================================

Confronta RIGA PER RIGA gli estratti conto UBS con Odoo per trovare TUTTE le differenze.

Autore: Business Analyst Agent
Data: 2024-11-16
"""

import xmlrpc.client
import csv
import json
from datetime import datetime, timedelta
from collections import defaultdict
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# CONFIGURAZIONE ODOO
ODOO_URL = 'https://lapadevadmin-lapa-v2-staging-2406-25408900.dev.odoo.com'
ODOO_DB = 'lapadevadmin-lapa-v2-staging-2406-25408900'
ODOO_USERNAME = 'paul@lapa.ch'
ODOO_PASSWORD = 'lapa201180'

# PATH ESTRATTI CONTO
UBS_CHF_FILES = [
    r"C:\Users\lapa\Downloads\CHIUSURA 2024\CHIUSURA 2024\UBS CHF\UBS CHF 1.1-31.3.2024.csv",
    r"C:\Users\lapa\Downloads\CHIUSURA 2024\CHIUSURA 2024\UBS CHF\UBS CHF 1.4-30.6.2024.csv",
    r"C:\Users\lapa\Downloads\CHIUSURA 2024\CHIUSURA 2024\UBS CHF\UBS CHF 1.7-30.9.2024.csv",
    r"C:\Users\lapa\Downloads\CHIUSURA 2024\CHIUSURA 2024\UBS CHF\UBS CHF 1.10-31.12.2024.csv"
]

UBS_EUR_FILES = [
    r"C:\Users\lapa\Downloads\CHIUSURA 2024\CHIUSURA 2024\UBS EUR\UBS EUR 1.1-30.6.2024.csv",
    r"C:\Users\lapa\Downloads\CHIUSURA 2024\CHIUSURA 2024\UBS EUR\UBS EUR 1.7-31.12.2024.csv"
]

# CONTI ODOO
KONTO_UBS_CHF = '1022'  # UBS 701J (CHF)
KONTO_UBS_EUR = '1023'  # UBS 760A (EUR)

class OdooClient:
    """Client per connettersi a Odoo"""

    def __init__(self, url, db, username, password):
        self.url = url
        self.db = db
        self.username = username
        self.password = password
        self.uid = None
        self.models = None

    def connect(self):
        """Connessione e autenticazione"""
        print(f"[*] Connessione a Odoo: {self.url}")
        common = xmlrpc.client.ServerProxy(f'{self.url}/xmlrpc/2/common')
        self.uid = common.authenticate(self.db, self.username, self.password, {})

        if not self.uid:
            raise Exception("[X] Autenticazione fallita!")

        print(f"[OK] Autenticato come UID: {self.uid}")
        self.models = xmlrpc.client.ServerProxy(f'{self.url}/xmlrpc/2/object')

    def get_account_moves(self, account_code, date_from, date_to):
        """Estrae movimenti contabili per un conto in un periodo"""

        # Trova l'account ID
        account_ids = self.models.execute_kw(
            self.db, self.uid, self.password,
            'account.account', 'search',
            [[('code', '=', account_code)]]
        )

        if not account_ids:
            print(f"[!] Conto {account_code} non trovato!")
            return []

        account_id = account_ids[0]

        # Estrae le righe contabili (account.move.line)
        domain = [
            ('account_id', '=', account_id),
            ('date', '>=', date_from),
            ('date', '<=', date_to),
            ('parent_state', '=', 'posted')  # Solo movimenti confermati
        ]

        fields = [
            'date', 'name', 'ref', 'debit', 'credit',
            'balance', 'move_id', 'partner_id', 'amount_currency'
        ]

        lines = self.models.execute_kw(
            self.db, self.uid, self.password,
            'account.move.line', 'search_read',
            [domain], {'fields': fields, 'order': 'date asc'}
        )

        return lines

class UBSParser:
    """Parser per estratti conto UBS in formato CSV"""

    @staticmethod
    def parse_date(date_str):
        """Converte data UBS (DD/MM/YYYY) in datetime"""
        try:
            return datetime.strptime(date_str, '%d/%m/%Y')
        except:
            return None

    @staticmethod
    def parse_amount(amount_str):
        """Converte importo UBS in Decimal"""
        if not amount_str:
            return Decimal('0')

        # Rimuove spazi e converte virgola in punto
        amount_str = amount_str.strip().replace("'", "").replace(",", ".")

        try:
            return Decimal(amount_str)
        except:
            return Decimal('0')

    @staticmethod
    def read_ubs_csv(file_path):
        """Legge file CSV UBS e estrae transazioni"""

        print(f"\n[FILE] Lettura file: {os.path.basename(file_path)}")

        transactions = []
        metadata = {}

        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f, delimiter=';')

            # Leggi metadati header
            for i, row in enumerate(reader):
                if i == 0:
                    metadata['account'] = row[1] if len(row) > 1 else ''
                elif i == 1:
                    metadata['iban'] = row[1] if len(row) > 1 else ''
                elif i == 2:
                    metadata['date_from'] = row[1] if len(row) > 1 else ''
                elif i == 3:
                    metadata['date_to'] = row[1] if len(row) > 1 else ''
                elif i == 4:
                    metadata['balance_start'] = UBSParser.parse_amount(row[1]) if len(row) > 1 else 0
                elif i == 5:
                    metadata['balance_end'] = UBSParser.parse_amount(row[1]) if len(row) > 1 else 0
                elif i == 6:
                    metadata['currency'] = row[1] if len(row) > 1 else ''
                elif i == 7:
                    metadata['transaction_count'] = row[1] if len(row) > 1 else ''
                elif i == 9:
                    # Header colonne transazioni
                    headers = row
                elif i > 9:
                    # Transazioni
                    if len(row) < 10:
                        continue

                    # Salta righe di dettaglio (senza data)
                    if not row[0]:
                        continue

                    date = UBSParser.parse_date(row[0])
                    if not date:
                        continue

                    debit = UBSParser.parse_amount(row[5])
                    credit = UBSParser.parse_amount(row[6])
                    balance = UBSParser.parse_amount(row[8])

                    # Importo netto (credit - debit)
                    amount = credit - debit

                    transaction = {
                        'date': date,
                        'booking_date': UBSParser.parse_date(row[2]) if row[2] else date,
                        'value_date': UBSParser.parse_date(row[3]) if row[3] else date,
                        'currency': row[4],
                        'debit': debit,
                        'credit': credit,
                        'amount': amount,
                        'balance': balance,
                        'transaction_nr': row[9],
                        'description': ' '.join([row[10], row[11], row[12]]).strip(),
                        'description1': row[10],
                        'description2': row[11],
                        'description3': row[12]
                    }

                    transactions.append(transaction)

        print(f"[OK] Transazioni lette: {len(transactions)}")
        print(f"     Saldo iniziale: {metadata['currency']} {metadata['balance_start']:,.2f}")
        print(f"     Saldo finale: {metadata['currency']} {metadata['balance_end']:,.2f}")

        return transactions, metadata

class BankReconciliation:
    """Engine di riconciliazione bancaria"""

    def __init__(self, odoo_client):
        self.odoo = odoo_client

    def match_transactions(self, bank_trans, odoo_trans, tolerance=0.01):
        """
        Matcha transazioni banca con Odoo usando:
        1. Data esatta + importo esatto
        2. Data +/- 2 giorni + importo esatto
        3. Data esatta + importo con tolleranza
        """

        matches = []
        unmatched_bank = []
        unmatched_odoo = list(odoo_trans)  # Copia

        for bt in bank_trans:
            matched = False

            # Prova match esatto (data + importo)
            for ot in unmatched_odoo[:]:
                odoo_date = datetime.strptime(ot['date'], '%Y-%m-%d')
                odoo_amount = Decimal(str(ot['debit'])) - Decimal(str(ot['credit']))

                # Match esatto
                if bt['date'].date() == odoo_date.date() and abs(bt['amount'] - odoo_amount) < Decimal(str(tolerance)):
                    matches.append({
                        'bank': bt,
                        'odoo': ot,
                        'match_type': 'exact'
                    })
                    unmatched_odoo.remove(ot)
                    matched = True
                    break

            if matched:
                continue

            # Prova match con data +/- 2 giorni
            for ot in unmatched_odoo[:]:
                odoo_date = datetime.strptime(ot['date'], '%Y-%m-%d')
                odoo_amount = Decimal(str(ot['debit'])) - Decimal(str(ot['credit']))

                date_diff = abs((bt['date'].date() - odoo_date.date()).days)

                if date_diff <= 2 and abs(bt['amount'] - odoo_amount) < Decimal(str(tolerance)):
                    matches.append({
                        'bank': bt,
                        'odoo': ot,
                        'match_type': 'fuzzy_date'
                    })
                    unmatched_odoo.remove(ot)
                    matched = True
                    break

            if not matched:
                unmatched_bank.append(bt)

        return matches, unmatched_bank, unmatched_odoo

    def reconcile_month(self, year, month, bank_trans, account_code):
        """Riconcilia un mese specifico"""

        # Filtra transazioni banca del mese
        month_bank = [
            t for t in bank_trans
            if t['date'].year == year and t['date'].month == month
        ]

        if not month_bank:
            return None

        # Estrae transazioni Odoo del mese
        date_from = f"{year}-{month:02d}-01"

        # Ultimo giorno del mese
        if month == 12:
            date_to = f"{year}-12-31"
        else:
            last_day = (datetime(year, month + 1, 1) - timedelta(days=1)).day
            date_to = f"{year}-{month:02d}-{last_day}"

        print(f"\n[MONTH] Riconciliazione {year}-{month:02d} (Conto {account_code})")
        print(f"        Periodo: {date_from} - {date_to}")

        odoo_moves = self.odoo.get_account_moves(account_code, date_from, date_to)

        print(f"        Transazioni banca: {len(month_bank)}")
        print(f"        Transazioni Odoo: {len(odoo_moves)}")

        # Match automatico
        matches, unmatched_bank, unmatched_odoo = self.match_transactions(
            month_bank, odoo_moves
        )

        # Calcola saldi
        bank_balance_change = sum(t['amount'] for t in month_bank)
        odoo_balance_change = sum(
            Decimal(str(m['debit'])) - Decimal(str(m['credit']))
            for m in odoo_moves
        )

        result = {
            'year': year,
            'month': month,
            'account_code': account_code,
            'bank_transactions': len(month_bank),
            'odoo_transactions': len(odoo_moves),
            'matched': len(matches),
            'unmatched_bank': len(unmatched_bank),
            'unmatched_odoo': len(unmatched_odoo),
            'bank_balance_change': float(bank_balance_change),
            'odoo_balance_change': float(odoo_balance_change),
            'balance_diff': float(bank_balance_change - odoo_balance_change),
            'matches_detail': matches,
            'unmatched_bank_detail': unmatched_bank,
            'unmatched_odoo_detail': unmatched_odoo
        }

        # Stampa summary
        print(f"\n        [OK] Matched: {len(matches)}")
        print(f"        [!] Non in Odoo: {len(unmatched_bank)}")
        print(f"        [!] Non in banca: {len(unmatched_odoo)}")
        print(f"        [$] Diff saldo: {result['balance_diff']:,.2f}")

        return result

class ExcelReporter:
    """Genera report Excel della riconciliazione"""

    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)  # Rimuove sheet default

    def add_summary_sheet(self, results):
        """Sheet 1: Summary annuale"""

        ws = self.wb.create_sheet("Summary 2024")

        # Header
        ws['A1'] = "RICONCILIAZIONE BANCARIA 2024 - SUMMARY"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:H1')

        # Headers tabella
        headers = ['Mese', 'Conto', 'Trans Banca', 'Trans Odoo', 'Matched',
                   'Non in Odoo', 'Non in Banca', 'Diff Saldo']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(3, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4472C4', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)

        # Dati
        row = 4
        total_matched = 0
        total_unmatched_bank = 0
        total_unmatched_odoo = 0

        for result in results:
            if not result:
                continue

            month_name = datetime(result['year'], result['month'], 1).strftime('%B %Y')

            ws.cell(row, 1, month_name)
            ws.cell(row, 2, result['account_code'])
            ws.cell(row, 3, result['bank_transactions'])
            ws.cell(row, 4, result['odoo_transactions'])
            ws.cell(row, 5, result['matched'])
            ws.cell(row, 6, result['unmatched_bank'])
            ws.cell(row, 7, result['unmatched_odoo'])
            ws.cell(row, 8, result['balance_diff'])

            # Formattazione
            if result['unmatched_bank'] > 0 or result['unmatched_odoo'] > 0:
                for col in range(1, 9):
                    ws.cell(row, col).fill = PatternFill(start_color='FFC7CE', fill_type='solid')

            total_matched += result['matched']
            total_unmatched_bank += result['unmatched_bank']
            total_unmatched_odoo += result['unmatched_odoo']

            row += 1

        # Totali
        row += 1
        ws.cell(row, 1, "TOTALE").font = Font(bold=True)
        ws.cell(row, 5, total_matched).font = Font(bold=True)
        ws.cell(row, 6, total_unmatched_bank).font = Font(bold=True)
        ws.cell(row, 7, total_unmatched_odoo).font = Font(bold=True)

        # Adjust column widths
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def add_detail_sheet(self, result):
        """Sheet per ogni mese con dettaglio discrepanze"""

        if not result:
            return

        month_name = datetime(result['year'], result['month'], 1).strftime('%Y-%m')
        sheet_name = f"{month_name} {result['account_code']}"

        ws = self.wb.create_sheet(sheet_name)

        # Title
        ws['A1'] = f"Riconciliazione {month_name} - Conto {result['account_code']}"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:F1')

        # Summary box
        ws['A3'] = "Transazioni banca:"
        ws['B3'] = result['bank_transactions']
        ws['A4'] = "Transazioni Odoo:"
        ws['B4'] = result['odoo_transactions']
        ws['A5'] = "Matched:"
        ws['B5'] = result['matched']
        ws['B5'].fill = PatternFill(start_color='C6EFCE', fill_type='solid')
        ws['A6'] = "Non in Odoo:"
        ws['B6'] = result['unmatched_bank']
        if result['unmatched_bank'] > 0:
            ws['B6'].fill = PatternFill(start_color='FFC7CE', fill_type='solid')
        ws['A7'] = "Non in banca:"
        ws['B7'] = result['unmatched_odoo']
        if result['unmatched_odoo'] > 0:
            ws['B7'].fill = PatternFill(start_color='FFC7CE', fill_type='solid')

        # Sezione: Transazioni NON in Odoo
        row = 10
        ws.cell(row, 1, "TRANSAZIONI BANCA NON IN ODOO").font = Font(bold=True, size=12)
        ws.cell(row, 1).fill = PatternFill(start_color='FFC7CE', fill_type='solid')
        ws.merge_cells(f'A{row}:F{row}')

        row += 1
        headers = ['Data', 'Importo', 'Descrizione', 'Trans Nr', 'Booking Date', 'Value Date']
        for col, header in enumerate(headers, 1):
            ws.cell(row, col, header).font = Font(bold=True)

        row += 1
        for trans in result['unmatched_bank_detail']:
            ws.cell(row, 1, trans['date'].strftime('%Y-%m-%d'))
            ws.cell(row, 2, float(trans['amount']))
            ws.cell(row, 3, trans['description'][:100])
            ws.cell(row, 4, trans['transaction_nr'])
            ws.cell(row, 5, trans['booking_date'].strftime('%Y-%m-%d'))
            ws.cell(row, 6, trans['value_date'].strftime('%Y-%m-%d'))
            row += 1

        # Sezione: Transazioni NON in banca
        row += 2
        ws.cell(row, 1, "TRANSAZIONI ODOO NON IN BANCA").font = Font(bold=True, size=12)
        ws.cell(row, 1).fill = PatternFill(start_color='FFC7CE', fill_type='solid')
        ws.merge_cells(f'A{row}:F{row}')

        row += 1
        headers = ['Data', 'Dare', 'Avere', 'Descrizione', 'Ref', 'Move ID']
        for col, header in enumerate(headers, 1):
            ws.cell(row, col, header).font = Font(bold=True)

        row += 1
        for move in result['unmatched_odoo_detail']:
            ws.cell(row, 1, move['date'])
            ws.cell(row, 2, float(move['debit']))
            ws.cell(row, 3, float(move['credit']))
            ws.cell(row, 4, str(move['name'])[:100])
            ws.cell(row, 5, move.get('ref', ''))
            ws.cell(row, 6, move['move_id'][0] if move.get('move_id') else '')
            row += 1

        # Adjust widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12

    def save(self):
        """Salva il file Excel"""
        self.wb.save(self.file_path)
        print(f"\n[SAVE] Report salvato: {self.file_path}")

def main():
    """Main execution"""

    print("="*80)
    print("RICONCILIAZIONE BANCARIA COMPLETA 2024 - MESE PER MESE")
    print("="*80)

    # 1. Connessione Odoo
    odoo = OdooClient(ODOO_URL, ODOO_DB, ODOO_USERNAME, ODOO_PASSWORD)
    odoo.connect()

    # 2. Leggi tutti gli estratti conto UBS CHF
    print("\n" + "="*80)
    print("[CHF] LETTURA ESTRATTI CONTO UBS CHF")
    print("="*80)

    all_chf_trans = []
    for file_path in UBS_CHF_FILES:
        trans, meta = UBSParser.read_ubs_csv(file_path)
        all_chf_trans.extend(trans)

    print(f"\n[OK] Totale transazioni CHF: {len(all_chf_trans)}")

    # 3. Leggi tutti gli estratti conto UBS EUR
    print("\n" + "="*80)
    print("[EUR] LETTURA ESTRATTI CONTO UBS EUR")
    print("="*80)

    all_eur_trans = []
    for file_path in UBS_EUR_FILES:
        trans, meta = UBSParser.read_ubs_csv(file_path)
        all_eur_trans.extend(trans)

    print(f"\n[OK] Totale transazioni EUR: {len(all_eur_trans)}")

    # 4. Riconciliazione mese per mese
    print("\n" + "="*80)
    print("[RECONCILE] RICONCILIAZIONE MESE PER MESE")
    print("="*80)

    reconciler = BankReconciliation(odoo)
    results = []

    # CHF - Gen-Dic 2024
    for month in range(1, 13):
        result = reconciler.reconcile_month(2024, month, all_chf_trans, KONTO_UBS_CHF)
        if result:
            results.append(result)

    # EUR - Gen-Dic 2024
    for month in range(1, 13):
        result = reconciler.reconcile_month(2024, month, all_eur_trans, KONTO_UBS_EUR)
        if result:
            results.append(result)

    # 5. Genera report Excel
    print("\n" + "="*80)
    print("[EXCEL] GENERAZIONE REPORT EXCEL")
    print("="*80)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    report_path = f"C:\\Users\\lapa\\Desktop\\Claude Code\\app-hub-platform\\scripts\\riconciliazione-bancaria-2024-{timestamp}.xlsx"

    reporter = ExcelReporter(report_path)
    reporter.add_summary_sheet(results)

    for result in results:
        reporter.add_detail_sheet(result)

    reporter.save()

    # 6. Salva anche JSON per analisi successive
    json_path = f"C:\\Users\\lapa\\Desktop\\Claude Code\\app-hub-platform\\scripts\\riconciliazione-bancaria-2024-{timestamp}.json"

    # Converti Decimal e datetime in formati JSON-serializable
    def json_serializer(obj):
        if isinstance(obj, Decimal):
            return float(obj)
        if isinstance(obj, datetime):
            return obj.isoformat()
        return str(obj)

    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=2, default=json_serializer)

    print(f"[SAVE] JSON salvato: {json_path}")

    # 7. Summary finale
    print("\n" + "="*80)
    print("[SUMMARY] SUMMARY FINALE")
    print("="*80)

    total_matched = sum(r['matched'] for r in results if r)
    total_unmatched_bank = sum(r['unmatched_bank'] for r in results if r)
    total_unmatched_odoo = sum(r['unmatched_odoo'] for r in results if r)

    print(f"\n[OK] Totale transazioni matched: {total_matched}")
    print(f"[!] Totale transazioni banca NON in Odoo: {total_unmatched_bank}")
    print(f"[!] Totale transazioni Odoo NON in banca: {total_unmatched_odoo}")

    if total_unmatched_bank > 0 or total_unmatched_odoo > 0:
        print("\n[WARN] ATTENZIONE: Trovate discrepanze!")
        print(f"       Controlla il report Excel: {report_path}")
    else:
        print("\n[SUCCESS] PERFETTO! Nessuna discrepanza trovata!")

    print("\n[DONE] Riconciliazione completata!")

if __name__ == '__main__':
    main()
