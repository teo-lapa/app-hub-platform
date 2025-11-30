#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
VERIFICA QUADRATURA GENERALE CONTABILITÀ 2024 - VERSIONE OTTIMIZZATA
====================================================================

Controlli fondamentali di partita doppia (versione veloce):
1. Trial Balance completo (DARE = AVERE)
2. Quadratura per mese
3. Conti sospesi non chiusi
4. Patrimonio Netto
5. Movimenti con contropartite mancanti (sample)

Output: VERIFICA-QUADRATURA-2024.xlsx
"""

import sys
import io
import xmlrpc.client
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Fix encoding per Windows
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# Configurazione Odoo
ODOO_URL = "https://lapadevadmin-lapa-v2-staging-2406-25408900.dev.odoo.com"
ODOO_DB = "lapadevadmin-lapa-v2-staging-2406-25408900"
ODOO_USERNAME = "paul@lapa.ch"
ODOO_PASSWORD = "lapa201180"

# Colori per Excel
COLOR_CRITICAL = "FFD9534F"
COLOR_HIGH = "FFF0AD4E"
COLOR_MEDIUM = "FFF9F871"
COLOR_LOW = "FF5BC0DE"
COLOR_OK = "FF5CB85C"
COLOR_HEADER = "FF34495E"

def connect_odoo():
    """Connessione a Odoo via XML-RPC"""
    print("Connessione a Odoo...")

    common = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/common')
    uid = common.authenticate(ODOO_DB, ODOO_USERNAME, ODOO_PASSWORD, {})

    if not uid:
        raise Exception("Autenticazione fallita!")

    print(f"Connesso come user ID: {uid}")

    models = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/object')
    return models, uid

def get_trial_balance(models, uid):
    """Estrae Trial Balance completo al 31.12.2024"""
    print("\n[STEP 1] Trial Balance al 31.12.2024...")

    # Fetch tutti i movimenti contabili 2024
    move_lines = models.execute_kw(
        ODOO_DB, uid, ODOO_PASSWORD,
        'account.move.line', 'search_read',
        [[
            ('date', '>=', '2024-01-01'),
            ('date', '<=', '2024-12-31'),
            ('parent_state', '=', 'posted')
        ]],
        {
            'fields': ['account_id', 'debit', 'credit', 'balance', 'date', 'move_id', 'name'],
            'limit': 150000
        }
    )

    print(f"   Trovati {len(move_lines)} movimenti contabili nel 2024")

    # Fetch dettagli conti
    account_ids = list(set([
        ml['account_id'][0]
        for ml in move_lines
        if ml.get('account_id') and ml['account_id'] is not False
    ]))

    print(f"   Trovati {len(account_ids)} conti unici")

    accounts = models.execute_kw(
        ODOO_DB, uid, ODOO_PASSWORD,
        'account.account', 'search_read',
        [[('id', 'in', account_ids)]],
        {'fields': ['id', 'code', 'name', 'account_type']}
    )

    accounts_map = {acc['id']: acc for acc in accounts}

    # Aggregazione per conto
    trial_balance = defaultdict(lambda: {
        'code': '',
        'name': '',
        'debit': 0.0,
        'credit': 0.0,
        'balance': 0.0,
        'account_type': ''
    })

    for ml in move_lines:
        if not ml.get('account_id') or ml['account_id'] is False:
            continue

        acc_id = ml['account_id'][0]
        acc = accounts_map.get(acc_id, {})

        trial_balance[acc_id]['code'] = acc.get('code', '???')
        trial_balance[acc_id]['name'] = acc.get('name', 'Unknown')
        trial_balance[acc_id]['account_type'] = acc.get('account_type', '')
        trial_balance[acc_id]['debit'] += ml['debit']
        trial_balance[acc_id]['credit'] += ml['credit']
        trial_balance[acc_id]['balance'] += ml['balance']

    # Ordina per codice conto
    trial_balance_sorted = sorted(
        trial_balance.items(),
        key=lambda x: x[1]['code']
    )

    return trial_balance_sorted, move_lines, accounts_map

def verify_balance_equation(trial_balance):
    """VERIFICA CRITICA: Somma(Dare) = Somma(Avere)"""
    print("\n[STEP 2] Verifica equazione contabile DARE = AVERE...")

    total_debit = sum(acc[1]['debit'] for acc in trial_balance)
    total_credit = sum(acc[1]['credit'] for acc in trial_balance)
    difference = abs(total_debit - total_credit)

    print(f"   Totale DARE:  {total_debit:,.2f} CHF")
    print(f"   Totale AVERE: {total_credit:,.2f} CHF")
    print(f"   Differenza:   {difference:,.2f} CHF")

    if difference < 0.01:
        print("   OK - QUADRATURA OK!")
        severity = "OK"
    elif difference < 1.00:
        print("   ATTENZIONE: Differenza < 1 CHF (arrotondamenti)")
        severity = "LOW"
    elif difference < 100.00:
        print("   WARNING: Differenza significativa!")
        severity = "MEDIUM"
    else:
        print("   CRITICAL: Contabilita NON quadra!")
        severity = "CRITICAL"

    return {
        'total_debit': total_debit,
        'total_credit': total_credit,
        'difference': difference,
        'severity': severity,
        'balanced': difference < 0.01
    }

def verify_monthly_balance(move_lines):
    """Verifica quadratura per ogni mese"""
    print("\n[STEP 3] Verifica quadratura mensile...")

    monthly = defaultdict(lambda: {'debit': 0.0, 'credit': 0.0})

    for ml in move_lines:
        month = ml['date'][:7]  # YYYY-MM
        monthly[month]['debit'] += ml['debit']
        monthly[month]['credit'] += ml['credit']

    monthly_results = []
    for month in sorted(monthly.keys()):
        data = monthly[month]
        diff = abs(data['debit'] - data['credit'])
        balanced = diff < 0.01

        severity = "OK" if balanced else ("LOW" if diff < 1 else ("MEDIUM" if diff < 100 else "CRITICAL"))

        monthly_results.append({
            'month': month,
            'debit': data['debit'],
            'credit': data['credit'],
            'difference': diff,
            'balanced': balanced,
            'severity': severity
        })

        status = "OK" if balanced else "ERR"
        print(f"   [{status}] {month}: Dare={data['debit']:,.2f}, Avere={data['credit']:,.2f}, Diff={diff:,.2f}")

    return monthly_results

def verify_suspense_accounts(trial_balance):
    """Verifica conti sospesi/transitivi"""
    print("\n[STEP 4] Verifica conti sospesi...")

    suspense_codes = ['1099', '10901', '1022', '1023', '1024']
    suspense_accounts = []

    for acc_id, acc_data in trial_balance:
        if acc_data['code'] in suspense_codes:
            balance = abs(acc_data['balance'])

            if balance < 0.01:
                severity = "OK"
                status = "CHIUSO"
            elif balance < 10:
                severity = "LOW"
                status = "Residuo minimo"
            elif balance < 100:
                severity = "MEDIUM"
                status = "Residuo significativo"
            else:
                severity = "CRITICAL"
                status = "NON CHIUSO"

            suspense_accounts.append({
                'code': acc_data['code'],
                'name': acc_data['name'],
                'debit': acc_data['debit'],
                'credit': acc_data['credit'],
                'balance': acc_data['balance'],
                'severity': severity,
                'status': status
            })

            print(f"   [{status}] {acc_data['code']} - {acc_data['name']}: {acc_data['balance']:,.2f} CHF")

    return suspense_accounts

def verify_equity(trial_balance):
    """Verifica Patrimonio Netto"""
    print("\n[STEP 5] Verifica Patrimonio Netto...")

    revenues = 0.0
    expenses = 0.0
    equity_2900 = 0.0

    for acc_id, acc_data in trial_balance:
        code = acc_data['code']
        balance = acc_data['balance']

        # Ricavi (3xxx, 4xxx)
        if code.startswith('3') or code.startswith('4'):
            revenues += abs(balance) if balance < 0 else -balance

        # Costi (5xxx, 6xxx, 7xxx)
        elif code.startswith('5') or code.startswith('6') or code.startswith('7'):
            expenses += abs(balance) if balance > 0 else -balance

        # Conto utile/perdita esercizio
        elif code == '2900':
            equity_2900 = balance

    calculated_profit = revenues - expenses
    difference = abs(calculated_profit - equity_2900)

    print(f"   Ricavi totali:        {revenues:,.2f} CHF")
    print(f"   Costi totali:         {expenses:,.2f} CHF")
    print(f"   Utile calcolato:      {calculated_profit:,.2f} CHF")
    print(f"   Conto 2900:           {equity_2900:,.2f} CHF")
    print(f"   Differenza:           {difference:,.2f} CHF")

    if difference < 0.01:
        print("   OK - Patrimonio Netto CORRETTO")
        severity = "OK"
    elif difference < 100:
        print("   ATTENZIONE: Differenza minore")
        severity = "LOW"
    else:
        print("   CRITICAL: Differenza significativa!")
        severity = "CRITICAL"

    return {
        'revenues': revenues,
        'expenses': expenses,
        'calculated_profit': calculated_profit,
        'account_2900': equity_2900,
        'difference': difference,
        'severity': severity
    }

def find_orphan_moves_sample(models, uid, move_lines):
    """Cerca movimenti senza contropartita (sample 1000 registrazioni)"""
    print("\n[STEP 6] Ricerca contropartite mancanti (sample)...")

    # Raggruppa per move_id
    moves_dict = defaultdict(list)
    for ml in move_lines[:20000]:  # Limita a prime 20000 righe per velocita
        moves_dict[ml['move_id'][0]].append(ml)

    orphan_moves = []

    for move_id, lines in moves_dict.items():
        # Una registrazione DEVE avere almeno 2 righe
        if len(lines) < 2:
            orphan_moves.append({
                'move_id': move_id,
                'move_name': lines[0]['move_id'][1] if lines else '???',
                'lines_count': len(lines),
                'reason': 'Una sola riga (manca contropartita)',
                'severity': 'CRITICAL'
            })

        # Verifica che ci sia almeno un dare e un avere
        has_debit = any(line['debit'] > 0 for line in lines)
        has_credit = any(line['credit'] > 0 for line in lines)

        if not (has_debit and has_credit):
            orphan_moves.append({
                'move_id': move_id,
                'move_name': lines[0]['move_id'][1] if lines else '???',
                'lines_count': len(lines),
                'reason': 'Mancano dare o avere',
                'severity': 'CRITICAL'
            })

    print(f"   Trovati {len(orphan_moves)} movimenti con problemi (su sample)")

    return orphan_moves

def create_excel_report(
    trial_balance,
    balance_check,
    monthly_balance,
    suspense_accounts,
    equity_check,
    orphan_moves
):
    """Genera report Excel completo"""
    print("\n[STEP 7] Generazione report Excel...")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Stili
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")

    severity_colors = {
        'OK': COLOR_OK,
        'LOW': COLOR_LOW,
        'MEDIUM': COLOR_MEDIUM,
        'HIGH': COLOR_HIGH,
        'CRITICAL': COLOR_CRITICAL
    }

    # SHEET 1: EXECUTIVE SUMMARY
    ws1 = wb.create_sheet("EXECUTIVE SUMMARY")
    ws1.append(["VERIFICA QUADRATURA GENERALE CONTABILITA 2024"])
    ws1.append([f"Generato il: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"])
    ws1.append([])

    ws1.append(["QUADRATURA GENERALE (DARE = AVERE)"])
    ws1.append(["Totale DARE", balance_check['total_debit'], "CHF"])
    ws1.append(["Totale AVERE", balance_check['total_credit'], "CHF"])
    ws1.append(["Differenza", balance_check['difference'], "CHF"])
    ws1.append(["Severity", balance_check['severity']])

    severity_row = ws1.max_row
    ws1[f'B{severity_row}'].fill = PatternFill(
        start_color=severity_colors.get(balance_check['severity'], COLOR_LOW),
        end_color=severity_colors.get(balance_check['severity'], COLOR_LOW),
        fill_type="solid"
    )

    ws1.append([])
    ws1.append(["PATRIMONIO NETTO"])
    ws1.append(["Ricavi", equity_check['revenues'], "CHF"])
    ws1.append(["Costi", equity_check['expenses'], "CHF"])
    ws1.append(["Utile/Perdita calcolato", equity_check['calculated_profit'], "CHF"])
    ws1.append(["Conto 2900", equity_check['account_2900'], "CHF"])
    ws1.append(["Differenza", equity_check['difference'], "CHF"])
    ws1.append(["Severity", equity_check['severity']])

    severity_row = ws1.max_row
    ws1[f'B{severity_row}'].fill = PatternFill(
        start_color=severity_colors.get(equity_check['severity'], COLOR_LOW),
        end_color=severity_colors.get(equity_check['severity'], COLOR_LOW),
        fill_type="solid"
    )

    ws1.append([])
    ws1.append(["PROBLEMI RILEVATI"])
    ws1.append(["Conti sospesi non chiusi", sum(1 for s in suspense_accounts if s['severity'] != 'OK')])
    ws1.append(["Movimenti con contropartite mancanti (sample)", len(orphan_moves)])

    # SHEET 2: TRIAL BALANCE
    ws2 = wb.create_sheet("TRIAL BALANCE")
    ws2.append(["Codice", "Nome Conto", "DARE", "AVERE", "Saldo", "Tipo Conto"])

    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for acc_id, acc_data in trial_balance:
        ws2.append([
            acc_data['code'],
            acc_data['name'],
            acc_data['debit'],
            acc_data['credit'],
            acc_data['balance'],
            acc_data['account_type']
        ])

    total_row = ws2.max_row + 1
    ws2.append([
        "TOTALE",
        "",
        balance_check['total_debit'],
        balance_check['total_credit'],
        balance_check['total_debit'] - balance_check['total_credit'],
        ""
    ])

    for cell in ws2[total_row]:
        cell.font = Font(bold=True)

    # SHEET 3: QUADRATURA MENSILE
    ws3 = wb.create_sheet("QUADRATURA MENSILE")
    ws3.append(["Mese", "DARE", "AVERE", "Differenza", "Quadra?", "Severity"])

    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill

    for month_data in monthly_balance:
        ws3.append([
            month_data['month'],
            month_data['debit'],
            month_data['credit'],
            month_data['difference'],
            "SI" if month_data['balanced'] else "NO",
            month_data['severity']
        ])

        row = ws3.max_row
        ws3[f'F{row}'].fill = PatternFill(
            start_color=severity_colors.get(month_data['severity'], COLOR_LOW),
            end_color=severity_colors.get(month_data['severity'], COLOR_LOW),
            fill_type="solid"
        )

    # SHEET 4: CONTI SOSPESI
    ws4 = wb.create_sheet("CONTI SOSPESI")
    ws4.append(["Codice", "Nome", "DARE", "AVERE", "Saldo", "Status", "Severity"])

    for cell in ws4[1]:
        cell.font = header_font
        cell.fill = header_fill

    for acc in suspense_accounts:
        ws4.append([
            acc['code'],
            acc['name'],
            acc['debit'],
            acc['credit'],
            acc['balance'],
            acc['status'],
            acc['severity']
        ])

        row = ws4.max_row
        ws4[f'G{row}'].fill = PatternFill(
            start_color=severity_colors.get(acc['severity'], COLOR_LOW),
            end_color=severity_colors.get(acc['severity'], COLOR_LOW),
            fill_type="solid"
        )

    # SHEET 5: CONTROPARTITE MANCANTI (SAMPLE)
    ws5 = wb.create_sheet("CONTROPARTITE MANCANTI")
    ws5.append(["Move ID", "Nome Move", "Righe Contabili", "Motivo", "Severity"])

    for cell in ws5[1]:
        cell.font = header_font
        cell.fill = header_fill

    for orphan in orphan_moves:
        ws5.append([
            orphan['move_id'],
            orphan['move_name'],
            orphan['lines_count'],
            orphan['reason'],
            orphan['severity']
        ])

        row = ws5.max_row
        ws5[f'E{row}'].fill = PatternFill(
            start_color=severity_colors.get(orphan['severity'], COLOR_LOW),
            end_color=severity_colors.get(orphan['severity'], COLOR_LOW),
            fill_type="solid"
        )

    # Auto-size colonne
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Salva
    filename = "VERIFICA-QUADRATURA-2024.xlsx"
    wb.save(filename)
    print(f"   Report salvato: {filename}")

    return filename

def main():
    """Main execution"""
    print("="*80)
    print("VERIFICA QUADRATURA GENERALE CONTABILITA 2024")
    print("="*80)

    try:
        # Connessione
        models, uid = connect_odoo()

        # 1. Trial Balance
        trial_balance, move_lines, accounts_map = get_trial_balance(models, uid)

        # 2. Verifica DARE = AVERE
        balance_check = verify_balance_equation(trial_balance)

        # 3. Quadratura mensile
        monthly_balance = verify_monthly_balance(move_lines)

        # 4. Conti sospesi
        suspense_accounts = verify_suspense_accounts(trial_balance)

        # 5. Patrimonio Netto
        equity_check = verify_equity(trial_balance)

        # 6. Contropartite mancanti (sample)
        orphan_moves = find_orphan_moves_sample(models, uid, move_lines)

        # 7. Report Excel
        filename = create_excel_report(
            trial_balance,
            balance_check,
            monthly_balance,
            suspense_accounts,
            equity_check,
            orphan_moves
        )

        # Summary finale
        print("\n" + "="*80)
        print("RIEPILOGO FINALE")
        print("="*80)
        print(f"Trial Balance: {len(trial_balance)} conti")
        print(f"Quadratura: {'OK' if balance_check['balanced'] else 'NON QUADRA'}")
        print(f"Mesi verificati: {len(monthly_balance)}")
        print(f"Conti sospesi non chiusi: {sum(1 for s in suspense_accounts if s['severity'] != 'OK')}")
        print(f"Patrimonio Netto: {'OK' if equity_check['severity'] == 'OK' else 'DA VERIFICARE'}")
        print(f"Contropartite mancanti (sample): {len(orphan_moves)}")
        print(f"\nReport: {filename}")
        print("="*80)

    except Exception as e:
        print(f"\nERRORE: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
