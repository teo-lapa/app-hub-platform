#!/usr/bin/env python3
"""
Esporta i dati UBS CHF 2024 in un file Excel formattato
"""

import json
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# File paths
JSON_FILE = r"C:\Users\lapa\Desktop\Claude Code\app-hub-platform\data-estratti\UBS-CHF-2024-CLEAN.json"
OUTPUT_FILE = r"C:\Users\lapa\Desktop\Claude Code\app-hub-platform\data-estratti\UBS-CHF-2024-REPORT.xlsx"

def load_data():
    """Load data from JSON"""
    with open(JSON_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def create_excel_report(data):
    """Create formatted Excel report"""
    if not OPENPYXL_AVAILABLE:
        print("[ERROR] openpyxl non installato. Installare con: pip install openpyxl")
        return

    # Create workbook
    wb = openpyxl.Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Create sheets
    create_summary_sheet(wb, data)
    create_monthly_sheet(wb, data)
    create_quarterly_sheet(wb, data)

    # Save
    wb.save(OUTPUT_FILE)
    print(f"[OK] File Excel creato: {OUTPUT_FILE}")

def create_summary_sheet(wb, data):
    """Create summary sheet"""
    ws = wb.create_sheet("Riepilogo", 0)

    # Title
    ws['A1'] = 'ANALISI ESTRATTI CONTO UBS CHF 2024'
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:D1')
    ws.row_dimensions[1].height = 30

    # Account info
    row = 3
    ws[f'A{row}'] = 'INFORMAZIONI CONTO'
    ws[f'A{row}'].font = Font(size=12, bold=True)
    ws.merge_cells(f'A{row}:D{row}')

    row += 1
    ws[f'A{row}'] = 'Numero Conto:'
    ws[f'B{row}'] = data['account']
    ws[f'A{row}'].font = Font(bold=True)

    row += 1
    ws[f'A{row}'] = 'IBAN:'
    ws[f'B{row}'] = data['iban']
    ws[f'A{row}'].font = Font(bold=True)

    row += 1
    ws[f'A{row}'] = 'Valuta:'
    ws[f'B{row}'] = data['currency']
    ws[f'A{row}'].font = Font(bold=True)

    row += 1
    ws[f'A{row}'] = 'Anno:'
    ws[f'B{row}'] = data['year']
    ws[f'A{row}'].font = Font(bold=True)

    # Balances
    row += 2
    ws[f'A{row}'] = 'SALDI ANNO 2024'
    ws[f'A{row}'].font = Font(size=12, bold=True)
    ws.merge_cells(f'A{row}:D{row}')

    row += 1
    headers = ['Descrizione', 'Importo (CHF)', '', '']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    row += 1
    ws[f'A{row}'] = 'Saldo 01/01/2024'
    ws[f'B{row}'] = data['saldo_inizio_anno']
    ws[f'B{row}'].number_format = '#,##0.00'

    row += 1
    ws[f'A{row}'] = 'Saldo 31/12/2024'
    ws[f'B{row}'] = data['saldo_fine_anno']
    ws[f'B{row}'].number_format = '#,##0.00'

    row += 1
    ws[f'A{row}'] = 'Variazione Anno'
    variation = data['saldo_fine_anno'] - data['saldo_inizio_anno']
    ws[f'B{row}'] = variation
    ws[f'B{row}'].number_format = '#,##0.00'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True, color='00FF00' if variation > 0 else 'FF0000')

    # Transactions
    row += 2
    ws[f'A{row}'] = 'TRANSAZIONI'
    ws[f'A{row}'].font = Font(size=12, bold=True)
    ws.merge_cells(f'A{row}:D{row}')

    row += 1
    ws[f'A{row}'] = 'Totale Transazioni Anno:'
    total_tx = sum(q['transactions'] for q in data['quarters'])
    ws[f'B{row}'] = total_tx
    ws[f'A{row}'].font = Font(bold=True)

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

def create_monthly_sheet(wb, data):
    """Create monthly balances sheet"""
    ws = wb.create_sheet("Saldi Mensili")

    # Title
    ws['A1'] = 'SALDI FINE MESE 2024'
    ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='4ECDC4', end_color='4ECDC4', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:D1')
    ws.row_dimensions[1].height = 25

    # Headers
    headers = ['Mese', 'Saldo Fine Mese (CHF)', 'Variazione', '% Variazione']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(3, col)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # Data
    row = 4
    prev_balance = data['saldo_inizio_anno']

    month_names = {
        '01': 'Gennaio', '02': 'Febbraio', '03': 'Marzo',
        '04': 'Aprile', '05': 'Maggio', '06': 'Giugno',
        '07': 'Luglio', '08': 'Agosto', '09': 'Settembre',
        '10': 'Ottobre', '11': 'Novembre', '12': 'Dicembre'
    }

    for month in sorted(data['monthly_balances'].keys()):
        balance = data['monthly_balances'][month]['ending_balance']
        variation = balance - prev_balance
        pct_variation = (variation / prev_balance * 100) if prev_balance != 0 else 0

        month_num = month.split('-')[1]
        month_name = f"{month_names[month_num]} {data['year']}"

        ws[f'A{row}'] = month_name
        ws[f'B{row}'] = balance
        ws[f'C{row}'] = variation
        ws[f'D{row}'] = pct_variation

        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'C{row}'].number_format = '#,##0.00'
        ws[f'D{row}'].number_format = '0.00%'

        # Color code variation
        if variation > 0:
            ws[f'C{row}'].font = Font(color='00FF00')
        elif variation < 0:
            ws[f'C{row}'].font = Font(color='FF0000')

        prev_balance = balance
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15

def create_quarterly_sheet(wb, data):
    """Create quarterly analysis sheet"""
    ws = wb.create_sheet("Analisi Trimestrale")

    # Title
    ws['A1'] = 'ANALISI TRIMESTRALE 2024'
    ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:G1')
    ws.row_dimensions[1].height = 25

    # Headers
    headers = ['Trimestre', 'Periodo', 'Saldo Iniziale', 'Saldo Finale', 'Variazione', 'Transazioni', '% Anno']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(3, col)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # Data
    total_transactions = sum(q['transactions'] for q in data['quarters'])
    row = 4

    for q in data['quarters']:
        variation = q['closing'] - q['opening']
        pct_of_year = (q['transactions'] / total_transactions * 100) if total_transactions > 0 else 0

        ws[f'A{row}'] = q['period']
        ws[f'B{row}'] = f"{q['date_from']} - {q['date_to']}"
        ws[f'C{row}'] = q['opening']
        ws[f'D{row}'] = q['closing']
        ws[f'E{row}'] = variation
        ws[f'F{row}'] = q['transactions']
        ws[f'G{row}'] = pct_of_year

        ws[f'C{row}'].number_format = '#,##0.00'
        ws[f'D{row}'].number_format = '#,##0.00'
        ws[f'E{row}'].number_format = '#,##0.00'
        ws[f'G{row}'].number_format = '0.00%'

        # Color code variation
        if variation > 0:
            ws[f'E{row}'].font = Font(color='00FF00', bold=True)
        elif variation < 0:
            ws[f'E{row}'].font = Font(color='FF0000', bold=True)

        row += 1

    # Totals row
    ws[f'A{row}'] = 'TOTALE'
    ws[f'F{row}'] = total_transactions
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'F{row}'].font = Font(bold=True)

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12

def main():
    print("=" * 80)
    print("EXPORT EXCEL - UBS CHF 2024")
    print("=" * 80)

    if not OPENPYXL_AVAILABLE:
        print("\n[ERROR] openpyxl non installato!")
        print("Installare con: pip install openpyxl")
        return

    print(f"\nCaricando dati da: {JSON_FILE}")
    data = load_data()

    print(f"Creando file Excel...")
    create_excel_report(data)

    print("\n" + "=" * 80)
    print("[OK] Export completato!")
    print("=" * 80)
    print(f"\nFile Excel: {OUTPUT_FILE}")

if __name__ == '__main__':
    main()
